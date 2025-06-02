import frappe
import requests
from datetime import date
import erpnext
import json
from frappe.utils import now
from frappe import throw,_
from frappe.utils import flt
from erpnext.setup.utils import get_exchange_rate
from frappe.utils import (
	add_days,
	ceil,
	cint,
	comma_and,
	flt,
	get_link_to_form,
	getdate,
	now_datetime,
	datetime,get_first_day,get_last_day,
	nowdate,
	today,
)
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days,date_diff
from datetime import date, datetime, timedelta
import datetime as dt
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, Union
from datetime import datetime
from onegene.mark_attendance import check_holiday
from frappe.utils.password import update_password
from frappe.utils.background_jobs import enqueue
import datetime as dt
from datetime import datetime, timedelta
from frappe.contacts.doctype.address.address import (
	get_address_display
)

import datetime
@frappe.whitelist()
def return_total_schedule(doc,method):
	# method to restrict the excess and low order scheduling
	total = frappe.db.sql(""" select `tabSales Order Schedule Item`.item_code, sum(`tabSales Order Schedule Item`.schedule_qty) as qty from `tabSales Order`
	left join `tabSales Order Schedule Item` on `tabSales Order`.name = `tabSales Order Schedule Item`.parent where `tabSales Order`.name = '%s' group by `tabSales Order Schedule Item`.item_code"""%(doc.name),as_dict = 1)

	item_total = frappe.db.sql(""" select `tabSales Order Item`.item_code, sum(`tabSales Order Item`.qty) as qty from `tabSales Order`
	left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent where `tabSales Order`.name = '%s' group by `tabSales Order Item`.item_code"""%(doc.name),as_dict = 1)
	for t in total:
		for i in item_total:
			if i.item_code == t.item_code:
				if t.qty > i.qty:
					frappe.throw(
						_(
							"Schedule Qty {2} is Greater than -  {0} for - {1}."
						).format(
							frappe.bold(i.qty),
							frappe.bold(i.item_code),
							frappe.bold(t.qty),
						)
					)
					frappe.validated = False
				if t.qty < i.qty:
					frappe.throw(
						_(
							"Schedule Qty {2} is Less than -  {0} for - {1}."
						).format(
							frappe.bold(i.qty),
							frappe.bold(i.item_code),
							frappe.bold(t.qty),
						)
					)
					frappe.validated = False

@frappe.whitelist()
def create_order_schedule_from_so(doc,method):
	if doc.customer_order_type == "Fixed" and not doc.custom_schedule_table:
		frappe.throw("Schedule not Created")
	if doc.customer_order_type == "Fixed" and doc.custom_schedule_table:
		for schedule in doc.custom_schedule_table:
			new_doc = frappe.new_doc('Sales Order Schedule') 
			new_doc.customer_code = doc.custom_customer_code
			new_doc.sales_order_number = doc.name
			new_doc.item_code = schedule.item_code
			new_doc.schedule_date = schedule.schedule_date
			new_doc.qty = schedule.schedule_qty
			for item in doc.items:
				if item.item_code == schedule.item_code:
					new_doc.child_name = schedule.name
					new_doc.schedule_amount = schedule.schedule_qty * item.rate
					new_doc.order_rate = item.rate
					new_doc.pending_qty = schedule.schedule_qty
					new_doc.pending_amount = schedule.schedule_qty * item.rate
			new_doc.save(ignore_permissions=True) 

@frappe.whitelist()
# method to delete the order schedule during so cancel
def cancel_order_schedule_on_so_cancel(doc,method):
	if doc.customer_order_type == "Fixed":
		exists = frappe.db.exists("Sales Order Schedule",{"sales_order_number":doc.name})
		if exists:
			os = frappe.db.get_all("Sales Order Schedule",{"sales_order_number":doc.name},'name')
			for o in os:
				print(o.name)
				delete_doc = frappe.get_doc('Sales Order Schedule',o.name)
				delete_doc.delete()

@frappe.whitelist()
def get_so_details(sales):
	dict_list = []
	so = frappe.get_doc("Sales Order",sales)
	for i in so.items:
		dict_list.append(frappe._dict({"name":i.name,"item_code":i.item_code,"pending_qty":i.qty,"bom":i.bom_no,"description": i.description,"warehouse":i.warehouse,"rate":i.rate,"amount":i.amount}))
	return dict_list

@frappe.whitelist()
def sample_check():
	item_code = "333QRJLA-EC03"
	sf = frappe.db.sql("""select `tabMaterial Request Item`.qty as qty from `tabMaterial Request`
		left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent
		where `tabMaterial Request Item`.item_code = '%s' and `tabMaterial Request`.docstatus != 2 and `tabMaterial Request`.transaction_date = CURDATE() """%(item_code),as_dict = 1)[0].qty or 0
	print(sf)

def get_exploded_items(bom, data, indent=0, qty=1):

	exploded_items = frappe.get_all(
		"BOM Item",
		filters={"parent": bom},
		fields=["qty", "bom_no", "qty", "item_code", "item_name", "description", "uom"],
	)

	for item in exploded_items:
		item["indent"] = indent
		data.append(
			{
				"item_code": item.item_code,
				"item_name": item.item_name,
				"indent": indent,
				"bom_level": indent,
				"bom": item.bom_no,
				"qty": item.qty * qty,
				"uom": item.uom,
				"description": item.description,
			}
		)
		if item.bom_no:
			get_exploded_items(item.bom_no, data, indent=indent + 1, qty=item.qty)

@frappe.whitelist()
def get_open_order(name, item_code, delivery_date, item_name, qty, rate, warehouse, amount):
	new_doc = frappe.new_doc('Open Order')
	new_doc.sales_order_number = name
	new_doc.set('open_order_table', [])
	new_doc.append("open_order_table", {
		"item_code": item_code,
		"delivery_date": delivery_date,
		"item_name": item_name,
		"rate": float(rate),
		"warehouse": warehouse,
		"amount": amount,
		"qty": qty,   
	})
	new_doc.save(ignore_permissions=True)
	return "ok"

@frappe.whitelist()
def create_order_schedule_from_so_for_open(item_code, schedule_date, schedule_qty, customer_code, name, rate):
	rate = float(rate)
	schedule_qty = float(schedule_qty)
	new_doc = frappe.new_doc('Sales Order Schedule') 
	new_doc.customer_code = customer_code
	new_doc.sales_order_number = name
	new_doc.item_code = item_code
	new_doc.schedule_date = schedule_date
	new_doc.qty = schedule_qty
	new_doc.schedule_amount = schedule_qty * rate
	new_doc.order_rate = rate
	new_doc.pending_qty = schedule_qty
	new_doc.pending_amount = schedule_qty * rate
	new_doc.save(ignore_permissions=True) 
	frappe.db.commit()



@frappe.whitelist()
def generate_production_plan():
	# create production plan report list based daily based on the scheduled date from order schedule
	from frappe.utils import getdate
	from datetime import datetime
	start_date = datetime.today().replace(day=1).date()
	work_order = frappe.db.sql("""
		SELECT item_code, item_name, item_group, SUM(pending_qty) AS qty
		FROM `tabSales Order Schedule`
		WHERE MONTH(schedule_date) = MONTH(CURRENT_DATE())
		GROUP BY item_code, item_name, item_group
	""", as_dict=1)
	for j in work_order:
		rej_allowance = frappe.get_value("Item",j.item_code,['rejection_allowance'])
		pack_size = frappe.get_value("Item",j.item_code,['pack_size'])
		fg_plan = frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['fg_kanban_qty']) or 0
		sfg_days = frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['sfg_days']) or 0
		today_plan = frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['today_production_plan']) or 0
		tent_plan_i= frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['tentative_plan_i']) or 0
		tent_plan_ii = frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['tentative_plan_ii']) or 0
		stock = frappe.db.sql(""" select sum(actual_qty) as actual_qty from `tabBin` where item_code = '%s' """%(j.item_code),as_dict = 1)[0]
		if not stock["actual_qty"]:
			stock["actual_qty"] = 0
		pos = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code,`tabDelivery Note Item`.qty as qty from `tabDelivery Note`
		left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
		where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.docstatus = 1 and `tabDelivery Note`.posting_date = CURDATE() """%(j.item_code),as_dict = 1)
		del_qty = 0
		if len(pos)>0:
			for l in pos:
				del_qty = l.qty
		delivery = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code,`tabDelivery Note Item`.qty as qty from `tabDelivery Note`
		left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
		where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.docstatus = 1 and `tabDelivery Note`.posting_date between '%s' and '%s' """%(j.item_code,start_date,today()),as_dict = 1)
		del_qty_as_on_date = 0
		if len(delivery)>0:
			for d in delivery:
				del_qty_as_on_date = d.qty
		produced = frappe.db.sql("""select `tabStock Entry Detail`.item_code as item_code,`tabStock Entry Detail`.qty as qty from `tabStock Entry`
		left join `tabStock Entry Detail` on `tabStock Entry`.name = `tabStock Entry Detail`.parent
		where `tabStock Entry Detail`.item_code = '%s' and `tabStock Entry`.docstatus = 1 and `tabStock Entry`.posting_date = CURDATE() and `tabStock Entry`.stock_entry_type = "Manufacture"  """%(j.item_code),as_dict = 1)
		prod = 0
		if len(produced)>0:
			for l in produced:
				prod = l.qty
		produced_as_on_date = frappe.db.sql("""select `tabStock Entry Detail`.item_code as item_code,`tabStock Entry Detail`.qty as qty from `tabStock Entry`
		left join `tabStock Entry Detail` on `tabStock Entry`.name = `tabStock Entry Detail`.parent
		where `tabStock Entry Detail`.item_code = '%s' and `tabStock Entry`.docstatus = 1 and `tabStock Entry`.posting_date between '%s' and '%s' and `tabStock Entry`.stock_entry_type = "Manufacture" """%(j.item_code,start_date,today()),as_dict = 1)
		pro_qty_as_on_date = 0
		if len(produced_as_on_date)>0:
			for d in produced_as_on_date:
				pro_qty_as_on_date = d.qty
		work_days = frappe.db.get_single_value("Production Plan Settings", "working_days")
		with_rej = (j.qty * (rej_allowance/100)) + j.qty
		per_day = j.qty / int(work_days)
		if pack_size > 0:
			cal = per_day/ pack_size
		total = ceil(cal) * pack_size
		today_balance = 0
		reqd_plan = 0
		balance = 0
		if with_rej and fg_plan:
			balance = (int(with_rej) + int(fg_plan))
			reqd_plan = (float(total) * float(sfg_days)) + float(fg_plan)
			today_balance = int(today_plan)-int(prod)
		td_balance = 0
		if today_balance > 0:
			td_balance = today_balance
		else:
			td_balance = 0
		exists = frappe.db.exists("Production Plan Report",{"date":today(),'item':j.item_code})
		if exists:
			doc = frappe.get_doc("Production Plan Report",{"date":today(),'item':j.item_code})
		else:
			doc = frappe.new_doc("Production Plan Report")
		doc.item = j.item_code
		doc.item_name = j.item_name
		doc.item_group = j.item_group
		doc.date = today()
		doc.rej_allowance = rej_allowance
		doc.monthly_schedule = with_rej
		doc.bin_qty = pack_size
		doc.per_day_plan = total
		doc.fg_kanban_qty = fg_plan
		doc.sfg_days = sfg_days
		doc.stock_qty = stock["actual_qty"]
		doc.delivered_qty = del_qty
		doc.del_as_on_yes = del_qty_as_on_date
		doc.produced_qty = prod
		doc.pro_as_on_yes = pro_qty_as_on_date
		doc.monthly_balance = balance
		doc.today_prod_plan = today_plan
		doc.today_balance = td_balance
		doc.required_plan = reqd_plan
		doc.tent_prod_plan_1 = tent_plan_i
		doc.tent_prod_plan_2 = tent_plan_ii
		doc.save(ignore_permissions=True)

@frappe.whitelist()
# error will be thrown when actice employee have relieving date
def inactive_employee(doc,method):
	if doc.status=="Active":
		if doc.relieving_date:
			throw(_("Please remove the relieving date for the Active Employee."))

@frappe.whitelist()
def list_all_raw_materials(order_schedule, scheduleqty):
	doc_list = []
	consolidated_items = {}

	self = frappe.get_doc("Sales Order Schedule", order_schedule)
	data = []
	bom_list = []

	bom = frappe.db.get_value("BOM", {'item': self.item_code}, ['name'])
	bom_list.append(frappe._dict({"bom": bom, "qty": scheduleqty}))

	for k in bom_list:
		get_exploded_items(k["bom"], data, k["qty"], bom_list)

	unique_items = {}
	for item in data:
		item_code = item['item_code']
		qty = item['qty']
		if item_code in unique_items:
			unique_items[item_code]['qty'] += qty
		else:
			unique_items[item_code] = item
	combined_items_list = list(unique_items.values())
	doc_list.append(combined_items_list)

	for i in doc_list:
		for h in i:
			item_code = h["item_code"]
			qty = h["qty"]
			if item_code in consolidated_items:
				consolidated_items[item_code] += qty
			else:
				consolidated_items[item_code] = qty
	return consolidated_items

def get_exploded_items(bom, data, qty, skip_list):
	exploded_items = frappe.get_all("BOM Item", filters={"parent": bom},
									fields=["qty", "bom_no", "item_code", "item_name", "description", "uom"])
	for item in exploded_items:
		item_code = item['item_code']
		if item_code in skip_list:
			continue
		item_qty = float(item['qty']) * float(qty)
		stock = frappe.db.get_value("Bin", {'item_code': item_code, 'warehouse': "SFS Store - O"},
									['actual_qty']) or 0
		to_order = item_qty - stock if item_qty > stock else 0
		data.append({
			"item_code": item_code,
			"qty": item_qty,
		})
		if item['bom_no']:
			get_exploded_items(item['bom_no'], data, qty=item_qty, skip_list=skip_list)


# The below two methods are called in MRP Test Report, Material Requirement Planning, Internal Material Request Plan
@frappe.whitelist()
def return_print(item_type,based_on):
	from frappe.utils import cstr, add_days, date_diff, getdate,today,gzip_decompress
	pr_name = frappe.db.get_value('Prepared Report', {'report_name': 'Material Requirements Planning','status':'Completed'}, 'name')
	attached_file_name = frappe.db.get_value("File",{"attached_to_doctype": 'Prepared Report',"attached_to_name": pr_name},"name",)
	attached_file = frappe.get_doc("File", attached_file_name)
	compressed_content = attached_file.get_content()
	uncompressed_content = gzip_decompress(compressed_content)
	dos = json.loads(uncompressed_content.decode("utf-8"))
	doc = frappe.new_doc("Material Request")
	doc.material_request_type = "Purchase"
	doc.transaction_date = frappe.utils.today()
	doc.schedule_date = frappe.utils.today()
	doc.set_warehouse = "Stores - O"
	if based_on == "Highlighted Rows":
		for i in dos['result']:
			if float(i['safety_stock']) > float(i['actual_stock_qty']):
				uom = frappe.db.get_value("Item",i['item_code'],'stock_uom')
				pps = frappe.db.sql("""select sum(actual_qty) as qty from `tabBin`
										where item_code = %s and warehouse != 'SFS Store - O' """, (i['item_code']), as_dict=1)[0].qty or 0
				sfs = frappe.db.sql("""select sum(actual_qty) as qty from `tabBin`
										where item_code = %s and warehouse = 'SFS Store - O' """, (i['item_code']), as_dict=1)[0].qty or 0

				if i['to_order'] > 0:
					doc.append("items", {
						'item_code': i['item_code'],
						'custom_item_type': i['item_type'],
						'schedule_date': frappe.utils.today(),
						'qty': i['to_order'],
						'custom_mr_qty': i['to_order'],
						'custom_total_req_qty': i['to_order'],
						'custom_current_req_qty': i['to_order'],
						'custom_stock_qty_copy': pps,
						'custom_shop_floor_stock': sfs,
						'custom_expected_date': i['expected_date'],
						# 'custom_today_req_qty': today_req,
						'uom': uom
					})
		doc.save()
		name = [
			"""<a href="/app/Form/Material Request/{0}">{1}</a>""".format(doc.name, doc.name)
		]
		frappe.msgprint(_("Material Request - {0} created").format(", ".join(name)))
	if based_on == "Item Type":
		for i in dos['result']:
			if i['item_type'] in item_type:
				uom = frappe.db.get_value("Item",i['item_code'],'stock_uom')
				pps = frappe.db.sql("""select sum(actual_qty) as qty from `tabBin`
										where item_code = %s and warehouse != 'SFS Store - O' """, (i['item_code']), as_dict=1)[0].qty or 0
				sfs = frappe.db.sql("""select sum(actual_qty) as qty from `tabBin`
										where item_code = %s and warehouse = 'SFS Store - O' """, (i['item_code']), as_dict=1)[0].qty or 0

				if i['to_order'] > 0:
					doc.append("items", {
						'item_code': i['item_code'],
						'custom_item_type': i['item_type'],
						'schedule_date': frappe.utils.today(),
						'qty': i['to_order'],
						'custom_mr_qty': i['to_order'],
						'custom_total_req_qty': i['to_order'],
						'custom_current_req_qty': i['to_order'],
						'custom_stock_qty_copy': pps,
						'custom_shop_floor_stock': sfs,
						'custom_expected_date': i['expected_date'],
						# 'custom_today_req_qty': today_req,
						'uom': uom
					})
		doc.save()
		name = [
			"""<a href="/app/Form/Material Request/{0}">{1}</a>""".format(doc.name, doc.name)
		]
		frappe.msgprint(_("Material Request - {0} created").format(", ".join(name)))

@frappe.whitelist()
def return_item_type():
	dict = []
	dict_list = []
	from frappe.utils import cstr, add_days, date_diff, getdate,today,gzip_decompress
	pr_name = frappe.db.get_value('Prepared Report', {'report_name': 'Material Requirements Planning','status':'Completed'}, 'name')
	attached_file_name = frappe.db.get_value("File",{"attached_to_doctype": 'Prepared Report',"attached_to_name": pr_name},"name",)
	attached_file = frappe.get_doc("File", attached_file_name)
	compressed_content = attached_file.get_content()
	uncompressed_content = gzip_decompress(compressed_content)
	dos = json.loads(uncompressed_content.decode("utf-8"))
	doc = frappe.new_doc("Material Request")
	doc.material_request_type = "Purchase"
	doc.transaction_date = frappe.utils.today()
	doc.schedule_date = frappe.utils.today()
	doc.set_warehouse = "Stores - O"
	for i in dos['result']:
		if i['item_type'] not in dict:
			dict.append(i['item_type'])
			dict_list.append(frappe._dict({'item_type':i['item_type']}))

	return dict_list

@frappe.whitelist()
def return_mr_details(mr):
	# method to set the expected delivery date in purchase order
	doc = frappe.get_doc("Material Request",mr)
	return doc.items

# The below two methods are called in MRP Test Report, Material Requirement Planning
@frappe.whitelist()
def stock_details_mpd_report(item):
	w_house = frappe.db.get_value("Warehouse",['name'])
	data = ''
	stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin where item_code = '%s' order by warehouse """%(item),as_dict=True)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan = 10><center>Stock Availability</center></th></tr>'
	data += '''
	<tr><td style="padding:1px;border: 1px solid black" colspan = 4><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black" colspan = 6>%s</td></tr>
	<tr><td style="padding:1px;border: 1px solid black" colspan = 4><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan = 6>%s</td></tr>'''%(item,frappe.db.get_value('Item',item,'item_name'))
	data += '''
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white"  colspan = 4><b>Warehouse</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white" colspan = 3><b>Stock Qty</b></td>
	</tr>'''
	i = 0
	for stock in stocks:
		if stock.warehouse != w_house:
			if stock.actual_qty > 0:
				data += '''<tr><td style="padding:1px;border: 1px solid black" colspan = 4 >%s</td><td style="padding:1px;border: 1px solid black" colspan = 3>%s</td></tr>'''%(stock.warehouse,stock.actual_qty)
	i += 1
	stock_qty = 0
	for stock in stocks:
		stock_qty += stock.actual_qty
	data += '''<tr><td style="background-color:#909e8a;padding:1px;border: 1px solid black;color:white;font-weight:bold" colspan = 4 >%s</td><td style="background-color:#909e8a;padding:1px;border: 1px solid black;color:white;font-weight:bold" colspan = 3>%s</td></tr>'''%("Total     ",stock_qty)
	data += '</table>'

	return data



@frappe.whitelist()
def stock_details_mpd(item,quantity):
	w_house = frappe.db.get_value("Warehouse",['name'])
	data = ''
	stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin where item_code = '%s' order by warehouse """%(item),as_dict=True)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan = 10><center>Stock Availability</center></th></tr>'
	data += '''
	<tr><td style="padding:1px;border: 1px solid black" colspan = 4><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black" colspan = 6>%s</td></tr>
	<tr><td style="padding:1px;border: 1px solid black" colspan = 4><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan = 6>%s</td></tr>'''%(item,frappe.db.get_value('Item',item,'item_name'))
	data += '''
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white"  colspan = 4><b>Warehouse</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white" colspan = 3><b>Stock Qty</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white" colspan = 3><b>Required Qty</b></td>

	</tr>'''
	req_qty = 0
	qty = frappe.get_doc("Material Planning Details",quantity)
	for q in qty.material_plan:
		req_qty += q.required_qty
	for stock in stocks:
		if stock.warehouse == w_house:
			if stock.actual_qty > 0:
				comp = frappe.get_value("Warehouse",stock.warehouse,['company'])
				data +=''' <tr><td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan = 4>%s</td><td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan = 3>%s</td><td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan = 3>%s</td></tr>'''%(stock.warehouse,stock.actual_qty,'')
	i = 0
	for stock in stocks:
		if stock.warehouse != w_house:
			if stock.actual_qty > 0:
				data += '''<tr><td style="padding:1px;border: 1px solid black" colspan = 4 >%s</td><td style="padding:1px;border: 1px solid black" colspan = 3>%s</td><td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan = 3>%s</td></tr>'''%(stock.warehouse,stock.actual_qty,"")
	i += 1
	stock_qty = 0
	for stock in stocks:
		stock_qty += stock.actual_qty
	data += '''<tr><td style="background-color:#909e8a;padding:1px;border: 1px solid black;color:white;font-weight:bold" colspan = 4 >%s</td><td style="background-color:#909e8a;padding:1px;border: 1px solid black;color:white;font-weight:bold" colspan = 3>%s</td><td style="background-color:#909e8a;color:white;padding:1px;border: 1px solid black;font-weight:bold" colspan = 3>%s</td></tr>'''%("Total     ",stock_qty,req_qty)
	data += '</table>'

	return data

@frappe.whitelist()
def previous_purchase(item_table):
	item_table = json.loads(item_table)
	data = []
	for item in item_table:
		try:
			item_name = frappe.get_value('Item',{'name':item['item_code']},"item_name")
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """%(item["item_code"]),as_dict=True)
			for po in pos:
				data.append([item['item_code'],item_name,po.qty])
		except:
			pass
	return data


@frappe.whitelist()
def previous_po_html(item_code):
	# method get the previous po details of the item on click
	data = ""
	item_name = frappe.get_value('Item',{'item_code':item_code},"item_name")
	pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order Item`.amount as amount,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
	left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
	where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by date"""%(item_code),as_dict=True)


	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan=6><center>Previous Purchase Order</center></th></tr>'
	data += '''
	<tr><td colspan =2 style="padding:1px;border: 1px solid black;width:300px" ><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black;width:200px" colspan =4>%s</td></tr>
	<tr><td colspan =2 style="padding:1px;border: 1px solid black" ><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan =4>%s</td></tr>

	<tr><td style="padding:1px;border: 1px solid black" colspan =1><b>Supplier Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>Previous Purchase Order</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Date</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Rate</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Quantity</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Amount</b>
	</td></tr>'''%(item_code,item_name)
	for po in pos:
		data += '''<tr>
			<td style="padding:1px;border: 1px solid black" colspan =1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td></tr>'''%(po.supplier,po.po,po.date,po.rate,po.qty,po.amount)

	data += '</table>'
	return data

# The below method is called in MRP Test Report, Material Requirement Planning
@frappe.whitelist()
def mpd_details(name):
	data = ""
	pos = frappe.db.sql("""select `tabMaterial Planning Item`.item_code,`tabMaterial Planning Item`.item_name,`tabMaterial Planning Item`.uom,`tabMaterial Planning Item`.order_schedule_date,sum(`tabMaterial Planning Item`.required_qty) as qty from `tabMaterial Planning Details`
		left join `tabMaterial Planning Item` on `tabMaterial Planning Details`.name = `tabMaterial Planning Item`.parent
		where `tabMaterial Planning Details`.name = '%s' group by `tabMaterial Planning Item`.order_schedule_date """%(name),as_dict = 1)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan=6><center>Order Schedule Details</center></th></tr>'
	data += '''
	<tr><td style="padding:1px;border: 1px solid black" colspan =1><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>UOM</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>Schedule Date</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>Quantity</b></td>
	</td></tr>'''
	for po in pos:
		data += '''<tr>
			<td style="padding:1px;border: 1px solid black" colspan =1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td></tr>'''%(po.item_code,po.item_name,po.uom,po.order_schedule_date,po.qty)
	data += '</table>'
	return data


@frappe.whitelist()
def list_raw_mat():
	qty = 120
	skip_list = []
	data = []
	bom = "BOM-742-HWFAB-002"
	exploded_items = frappe.get_all("BOM Item", filters={"parent": bom},fields=["qty", "bom_no as bom", "item_code", "item_name", "description", "uom"])
	for item in exploded_items:
		item_code = item['item_code']
		if item_code in skip_list:
			continue
		item_qty = flt(item['qty']) * qty
		data.append({"item_code": item_code,"item_name": item['item_name'],"bom": item['bom'],"uom": item['uom'],"qty": item_qty,"description": item['description']})
	frappe.errprint(data)


@frappe.whitelist()
# method to fetch the operation item list linked to the bom and append in the table
def get_bom_details(bo, child):
	dict_list = []
	seen_items = set()

	so = frappe.get_doc("BOM", bo)
	op = frappe.db.get_all("Operation Item List", {"operation_name": child, "document_name": bo}, ["*"])

	if op:
		checked_row = 0
		for j in op:
			checked_row = j.selected_field
			if j.item not in seen_items:
				dict_list.append(frappe._dict({"check_box": 1, "name": checked_row, "item_code": j.item, "req_tot_qty": j.req_tot_qty, "uom": j.uom}))
				seen_items.add(j.item)

	for i in so.items:
		if i.item_code not in seen_items:
			dict_list.append(frappe._dict({"item_code": i.item_code, "req_tot_qty": i.qty, "uom": i.uom}))
			seen_items.add(i.item_code)

	return dict_list

@frappe.whitelist()
# method to create operation item list on button click from BOM
def table_multiselect(docs,item,item_code,child,uom,req_tot_qty):
	op = frappe.db.get_value("Operation Item List",{"document_name":docs,"item":item_code,"operation_name":child},["name"])
	if not op:
		bom_child = frappe.new_doc("Operation Item List")
		bom_child.document_name = docs
		bom_child.item = item_code
		bom_child.operation_name = child
		bom_child.selected_field = item
		bom_child.req_tot_qty = req_tot_qty
		bom_child.uom = uom
		bom_child.save()

@frappe.whitelist()
def bday_allocate():
	# create additional salary automatically when employee have birthday on that month
	employee_query = """
	SELECT *
	FROM `tabEmployee`
	WHERE
		status = 'Active'
		AND employee_category IN ('Staff', 'Operator', 'Sub Staff')
		AND MONTH(date_of_birth) = MONTH(CURDATE())
		AND date_of_joining < CURDATE()
	"""
	employee = frappe.db.sql(employee_query, as_dict=True)
	pay =  get_first_day(nowdate())
	for emp in employee:
		if frappe.db.exists("Salary Structure Assignment",{'employee':emp.name,'docstatus':1}):
			if not frappe.db.exists('Additional Salary',{'employee':emp.name,'payroll_date':pay,'salary_component':"Birthday Allowance",'docstatus':('!=',2)}):
				bday_amt = frappe.new_doc("Additional Salary")
				bday_amt.employee = emp.name
				bday_amt.payroll_date = pay
				bday_amt.company = emp.company
				bday_amt.salary_component = "Birthday Allowance"
				bday_amt.currency = "INR"
				bday_amt.amount = 1000
				bday_amt.save(ignore_permissions = True)
				bday_amt.submit()




@frappe.whitelist()
def overtime_hours(doc,method):
	ot_hours=frappe.db.sql("""select sum(custom_overtime_hours) from `tabAttendance` where employee = '%s' and attendance_date between '%s' and '%s'"""%(doc.employee,doc.start_date,doc.end_date),as_dict=True)[0]
	doc.custom_overtime_hours = ot_hours['sum(custom_overtime_hours)']

@frappe.whitelist()
# method to set the fixed salary in salary slip
def fixed_salary(doc,method):
	if doc.salary_structure!='Operators' and doc.salary_structure!='Apprentice':
		base_amount=frappe.db.get_value("Salary Structure Assignment",{'employee':doc.employee,'docstatus':1},['base'], order_by="modified desc")
		earned_basic=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Basic"},["amount"]) or 0
		da=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Dearness Allowance"},["amount"]) or 0
		hra=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"House Rent Allowance"},["amount"]) or 0
		wa=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Washing Allowance"},["amount"]) or 0
		ca=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Conveyance Allowance"},["amount"]) or 0
		ea=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Education Allowance"},["amount"]) or 0
		pa=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Performance Allowance"},["amount"]) or 0
		sa=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Special Allowance"},["amount"]) or 0
		stipend=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Stipend"},["amount"]) or 0
		att_inc=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Attendance Incentive"},["amount"]) or 0
		basic_da=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Basic & DA"},["amount"]) or 0
		lta=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Leave Travel Allowance"},["amount"]) or 0
		mnc=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Medical & Conveyance Allowance"},["amount"]) or 0
		sp=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Special Pay"},["amount"]) or 0
		frappe.errprint('base_amount')
		frappe.errprint(base_amount)
		if doc.designation=='Asst General Manager':
			basic=base_amount*0.4
			hra= base_amount*0.25
			pa=base_amount * 0.10
		else:
			basic=base_amount*0.5
			hra= base_amount*0.2
			pa=base_amount * 0.05
		lta=base_amount * 0.05
		mnc=base_amount * 0.03
		sa=base_amount * 0.17
		total = basic+hra+pa+lta+mnc+sa
		doc.custom_basic = basic
		doc.custom_dearness_allowance = da
		doc.custom_house_rent_allowance = hra
		doc.custom_performance_allowance = pa
		doc.custom_special_allowance = sa
		doc.custom_basic_da = basic
		doc.custom_leave_travel_allowance = lta
		doc.custom_medical_conveyance_allowance = mnc
		doc.custom_total_fixed_amount = round(total)
		doc.save(ignore_permissions = True)
		frappe.db.commit()
		
	if doc.salary_structure=='Operators':
		base=frappe.db.get_value("Salary Structure Assignment",{'employee':doc.employee,'docstatus':1},['base'], order_by="modified desc")
		basic=base*0.25
		da=base * 0.25
		hra=base * 0.2
		wa=base * 0.05
		ca=base * 0.1
		ea=base * 0.1
		pa=base * 0.05
		epf=basic+da
		frappe.db.set_value("Salary Slip",doc.name,'custom_basic',basic)
		frappe.db.set_value("Salary Slip",doc.name,'custom_dearness_allowance',da)
		frappe.db.set_value("Salary Slip",doc.name,'custom_house_rent_allowance',hra)
		frappe.db.set_value("Salary Slip",doc.name,'custom_washing_allowance',wa)
		frappe.db.set_value("Salary Slip",doc.name,'custom_conveyance_allowance',ca)
		frappe.db.set_value("Salary Slip",doc.name,'custom_education_allowance',ea)
		frappe.db.set_value("Salary Slip",doc.name,'custom_performance_allowance',pa)
		# frappe.db.set_value("Salary Slip",doc.name,'custom_attendance_incentive',1000)
	if doc.salary_structure=='Apprentice':
		base=frappe.db.get_value("Salary Structure Assignment",{'employee':doc.employee,'docstatus':1},['base'], order_by="modified desc")
		basic=base*0.5
		da=base * 0.5
		frappe.db.set_value("Salary Slip",doc.name,'custom_basic',basic)
		frappe.db.set_value("Salary Slip",doc.name,'custom_dearness_allowance',da)
		# frappe.db.set_value("Salary Slip",doc.name,'custom_attendance_incentive',1000)
		frappe.db.set_value("Salary Slip",doc.name,'custom_total_fixed_amount',basic+da)
	if doc.salary_structure=='Trainee':
		base=frappe.db.get_value("Salary Structure Assignment",{'employee':doc.employee,'docstatus':1},['base'], order_by="modified desc")
		basic=base*0.5
		da=base * 0.5
		frappe.db.set_value("Salary Slip",doc.name,'custom_basic',basic)
		frappe.db.set_value("Salary Slip",doc.name,'custom_dearness_allowance',da)
		# frappe.db.set_value("Salary Slip",doc.name,'custom_attendance_incentive',1000)
		frappe.db.set_value("Salary Slip",doc.name,'custom_total_fixed_amount',basic+da)

@frappe.whitelist()
def sick_leave_allocation():
	# sl will be added 0.5 every month
	import datetime
	from datetime import date
	today = date.today()
	year_start_date = date(today.year, 1, 1)
	year_end_date = date(today.year, 12, 31)
	employees=frappe.db.get_all("Employee",{"Status":"Active"},['*'])
	for emp in employees:
		# if emp.employee_category=='Apprentice':
		if emp.employee_category=='Staff' or emp.employee_category=='Sub Staff' or emp.employee_category=='Operator':
			frappe.errprint(emp.name)
			la=frappe.db.exists("Leave Allocation",{"employee":emp.name,"leave_type":"Sick Leave","from_date":year_start_date,"to_date":year_end_date,'docstatus':1})
			if la:
				leave_all=frappe.get_doc("Leave Allocation",la)
				leave_all.new_leaves_allocated +=0.5
				leave_all.total_leaves_allocated +=0.5
				leave_all.save(ignore_permissions=True)
				leave_all.submit()

			else:
				leave_all=frappe.new_doc("Leave Allocation")
				leave_all.employee=emp.name
				leave_all.leave_type="Sick Leave"
				leave_all.from_date=year_start_date
				leave_all.to_date=year_end_date
				leave_all.new_leaves_allocated=0.5
				leave_all.total_leaves_allocated =0.5
				# leave_all.carry_forward=1
				leave_all.save(ignore_permissions=True)
				leave_all.submit()

def update_leave_policy():
	# earned leave will be allocated automatically based on the present and half days
	pre_year = date.today().year - 1
	start_of_year = date(pre_year, 1, 1)
	end_of_year = date(pre_year, 12, 31)
	current_year = date.today().year
	start = date(current_year, 1, 1)
	end = date(current_year, 12, 31)
	leave = frappe.get_all("Leave Policy Detail", ["leave_type", "annual_allocation"])
	for i in leave:
		if i.leave_type =="Earned Leave":
			employees = frappe.get_all("Employee",{"status": "Active",'employee_category':('!=','Contractor')},["name","company"])
			for emp in employees:
				present = frappe.db.count("Attendance",{"employee":emp.name,"status":"Present","attendance_date": ["between", [start_of_year, end_of_year]]})
				half_day = frappe.db.count("Attendance",{"employee":emp.name,"status":"Half Day","attendance_date": ["between", [start_of_year, end_of_year]]})
				half = half_day/2
				attendance = present + half
				earned_leave = round(attendance /20)
				if earned_leave:
					allow = frappe.new_doc("Leave Allocation")
					allow.employee = emp.name
					allow.company = emp.company
					allow.leave_type = "Earned Leave"
					allow.from_date = start
					allow.to_date = end
					allow.new_leaves_allocated = earned_leave
					allow.total_leaves_allocated = earned_leave
					allow.save(ignore_permissions=True)
					allow.submit()
	frappe.db.commit()



@frappe.whitelist()
def update_shift(employee,from_date,to_date):
	shift_3 = frappe.db.count("Attendance",{"employee":employee,"attendance_date": ["between", [from_date, to_date]],"status":"Present","shift":"3"})
	shift_3_half = frappe.db.count("Attendance",{"employee":employee,"attendance_date": ["between", [from_date, to_date]],"status":"Half Day","shift":"3"})
	half_3 = shift_3_half/2
	shift3 = shift_3 + half_3
	shift_5 = frappe.db.count("Attendance",{"employee":employee,"attendance_date": ["between", [from_date, to_date]],"status":"Present","shift":"5"})
	shift_5_half = frappe.db.count("Attendance",{"employee":employee,"attendance_date": ["between", [from_date, to_date]],"status":"Half Day","shift":"5"})
	half_5 = shift_5_half/2
	shift5 = shift_5 + half_5
	shift = shift3 + shift5
	return shift


from frappe.utils import cstr, cint, getdate,get_first_day, get_last_day, today, time_diff_in_hours
@frappe.whitelist()
# return the time diff for the permission time
def att_req_hours(f_time,t_time,custom_session,custom_shift):
	if custom_session == "Flexible":
		if f_time and t_time:
			# frappe.errprint("hlo")
			time_diff = time_diff_in_hours(t_time,f_time)
			return time_diff
	elif custom_session == "Full Day":
		return "8"
	else :
		return "4"

@frappe.whitelist()
def od_hours_update(doc, method):
	# update the attendance status based on the sessions from od
	if doc.workflow_state=='Approved':
		dates = get_dates(doc.from_date, doc.to_date)
		for date in dates:
			# update attendance with present when session is full day
			if doc.reason == "On Duty" and doc.custom_session == "Full Day":
				if frappe.db.exists("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)}):
					att = frappe.get_doc("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)})
				else:
					att = frappe.new_doc("Attendance")
					att.employee = doc.employee
				att.company = doc.company
				att.status = "Present"
				att.working_hours = 8
				att.attendance_request = doc.name
				att.save(ignore_permissions=True)
				att.submit()
				frappe.db.commit()
			# update attendance status based on session and existing working hours
			if doc.reason == "On Duty" and doc.custom_session in ["First Half", "Second Half"]:
				if frappe.db.exists("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)}):
					att = frappe.get_doc("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)})
					if att.working_hours >= 4:
						att.working_hours += 4
						att.status = "Present"
					else:
						att.working_hours += 4
						att.status = "Half Day"
					att.company = doc.company
				else:
					att = frappe.new_doc("Attendance")
					att.employee = doc.employee
					att.working_hours = 4
					att.company = doc.company
					att.status = "Half Day"
				att.attendance_request = doc.name
				att.save(ignore_permissions=True)
				att.submit()
				frappe.db.commit()
			if doc.reason == "On Duty" and doc.custom_session == "Flexible":
				if frappe.db.exists("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)}):
					att = frappe.get_doc("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)})
					if att.in_time and att.out_time:
						st = datetime.strptime(str(doc.custom_from_time), '%H:%M:%S').time()
						start_time = dt.datetime.combine(att.attendance_date,st)
						if att.in_time > start_time :
							att.in_time = start_time
						et = datetime.strptime(str(doc.custom_to_time), '%H:%M:%S').time()
						end_time = dt.datetime.combine(att.attendance_date,et)
						if att.out_time < end_time :
							att.out_time = end_time
						att.save(ignore_permissions=True)
						frappe.db.commit()
						
					

def get_dates(from_date,to_date):
	# method to get the dates from the given range
	no_of_days = date_diff(add_days(to_date, 1), from_date)
	dates = [add_days(from_date, i) for i in range(0, no_of_days)]
	return dates

@frappe.whitelist()
def update_birthday_alowance(doc,method):
	# add birthday allowance for left employee , birthday is on same as relieving month
	if doc.status == "Left":
		if doc.date_of_birth > doc.relieving_date:
			if doc.date_of_birth.month == doc.relieving_date.month:
				first_day = get_first_day(doc.relieving_date)
				if frappe.db.exists("Additional Salary", {'employee': doc.name, 'salary_component': "Birthday Allowance", 'payroll_date': first_day, 'docstatus': 1}):
					ad = frappe.get_doc("Additional Salary", {'employee': doc.name, 'salary_component': "Birthday Allowance", 'payroll_date': first_day, 'docstatus': 1})
					ad.update({
						'docstatus': 2
					})
					ad.save()


@frappe.whitelist()
def create_lwf():
	# add lwf deduction amount 20 rupees by default on december.
	def is_december_1(date_to_check):
		return date_to_check.month == 12 and date_to_check.day == 1
	employee_query = """
	SELECT *
	FROM `tabEmployee`
	WHERE
		status = 'Active'  """
	employee = frappe.db.sql(employee_query, as_dict=True)
	date_to_check = date.today()
	if is_december_1(date_to_check):
		print("The date is December 1st.")
		for emp in employee:
			if frappe.db.exists("Salary Structure Assignment", {'employee': emp.name, 'docstatus': 1}):
				if not frappe.db.exists('Additional Salary', {'employee': emp.name, 'payroll_date': date_to_check, 'salary_component': "Labour Welfare Fund", 'docstatus': ('!=', 2)}):
					lwf = frappe.new_doc("Additional Salary")
					lwf.employee = emp.name
					lwf.payroll_date = date_to_check
					lwf.company = emp.company
					lwf.salary_component = "Labour Welfare Fund"
					lwf.currency = "INR"
					lwf.amount = 20
					lwf.save(ignore_permissions=True)
					lwf.submit()
	else:
		print("The date is not December 1st.")

@frappe.whitelist()
def renamed_doc(doc,method):
	# method used to rename employee
	name = doc.name
	employee_number = doc.employee_number
	emp = frappe.get_doc("Employee",name)
	emps=frappe.get_all("Employee",{"status":"Active"},['*'])
	for i in emps:
		if emp.employee_number == employee_number:
			pass
		elif i.employee_number == employee_number:
			frappe.throw(f"Employee Number already exists for {i.name}")
		else:
			frappe.db.set_value("Employee",name,"employee_number",employee_number)
			frappe.rename_doc("Employee", name, employee_number, force=1)


@frappe.whitelist(allow_guest=True)
# code for live attendance page in home
def get_live_attendance():
	nowtime = datetime.now()
	att_details = {}
	att_details['nowtime'] = datetime.strftime(nowtime, '%d-%m-%Y %H:%M:%S')
	max_out = datetime.strptime('06:30', '%H:%M').time()

	if nowtime.time() > max_out:
		date1 = nowtime.date()
	else:
		date1 = (nowtime - timedelta(days=1)).date()

	staff_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Staff", "Sub Staff", "Director")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['staff_count'] = staff_count[0].count if staff_count else 0
	trainee_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Trainee")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['trainee_count'] = trainee_count[0].count if trainee_count else 0
	ops_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Operator")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['ops_count'] = ops_count[0].count if ops_count else 0
	aps_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Apprentice")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['aps_count'] = aps_count[0].count if aps_count else 0
	cl_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Contractor")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['cl_count'] = cl_count[0].count if cl_count else 0
	tot_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['tot_count'] = tot_count[0].count if tot_count else 0
	return att_details

@frappe.whitelist()
def update_leave_ledger():
	leave_updates = [
	"""update `tabAttendance` set 
		late_entry = 0,
		early_exit = 0,
		custom_late_entry_time = NULL,
		custom_early_out_time = NULL 
	where status = "Half Day" and attendance_date between "2024-03-01" and "2024-05-31"
	"""
]

	for query in leave_updates:
		leave = frappe.db.sql(query, as_dict=True)

@frappe.whitelist()
def check_pf_type(name):
	# method to return print based on the PF eligility
	if frappe.db.exists("Salary Detail",{"parent":name,"salary_component":"Provident Fund"}):
		return "With PF"
	else:
		return "Without PF"


@frappe.whitelist()
def mark_disable(doc,method):
	# disable user if employee status is left
	if doc.status=='Left':
		frappe.db.set_value("User",doc.user_id,"enabled",0)  


@frappe.whitelist()
def update_role(id):
	# If employee category is staff/sub staff then it enables the below role in user list.
	usr=frappe.get_doc("User",id)
	usr.append("roles",{
		"role":"Staff/Sub Staff"
	})
	usr.save(ignore_permissions=True)
	frappe.db.commit()

@frappe.whitelist()
def remove_system_manager_role(doc,method):
	# remove system manager from the roles after creating user document
	usr=frappe.get_doc("User",doc.name)
	usr.remove_roles("System Manager")
	usr.save(ignore_permissions=True)
	frappe.db.commit()

@frappe.whitelist()
def create_user_id(doc,method):
	# method to create user after the employee mis created against the employee
	user_id=doc.name.lower()+'@onegeneindia.in'
	password = "wonjin@321"
	if frappe.db.exists("User",{"email":user_id}):
		frappe.throw("User ID already exists")
	else:
		user=frappe.new_doc("User")
		user.first_name=doc.first_name
		user.middle_name=doc.middle_name
		user.last_name=doc.last_name
		user.username=doc.employee
		user.full_name=doc.employee_name
		user.email=user_id
		
		user.save(ignore_permissions=True)
		frappe.db.commit()
		from frappe.utils.password import update_password
		update_password(user=user_id, pwd=password)
		frappe.db.set_value("Employee",doc.name,'user_id',user_id)
		frappe.db.set_value('Employee',doc.name, 'create_user_permission', 1)


@frappe.whitelist()
def get_deleted_automatically():
	# delete the night shift planning document when the overtime is not done
	yesterday = add_days(today(), -1)
	planning = frappe.db.exists("Night Shift Auditors Planning List", {'attendance_date': yesterday})
	if planning:
		attendance_exists = frappe.db.exists("Attendance", {'employee': planning.emp, 'attendance_date': yesterday, 'docstatus': ('!=', 2)})
		if attendance_exists:
			attendance = frappe.get_doc("Attendance", {'employee': planning.emp, 'attendance_date': yesterday, 'docstatus': ('!=', 2)})
			date1 = dt.datetime.strptime(yesterday, "%Y-%m-%d").date()
			shift_end_time = datetime.strptime("05:00:00", '%H:%M:%S').time()
			start_time = dt.datetime.combine(add_days(date1,1), shift_end_time)
			if attendance.out_time :
				if attendance.out_time > start_time:
					status = "Eligible"
				else:
					status = "Not-Eligible"
			else:
				status = "Not-Eligible"
			if status == "Not-Eligible":
				frappe.delete_doc("Night Shift Auditors Planning List", planning.name, ignore_permissions=True)



@frappe.whitelist()
# method returns the live attendance shift and department wise
def get_data_system(date):
	data =""
	shift=frappe.get_all("Shift Type",{'name':('!=',"4")},['*'],order_by='name ASC')
	shift2=4
	for i in shift:
		shift2+=1
	ec1=0
	ec_count=frappe.get_all("Employee Category",{'name':('not in',['Sub Staff','Director','Trainee'])},['*'])
	for i in ec_count:
		ec1 +=1 
	data = "<table class='table table-bordered=1'>"
	data += "<tr><td colspan ={}  style='border: 1px solid black;background-color:#f6d992;text-align:center'><b>Live Attendance</b></td><td colspan ={} style='border: 1px solid black;background-color:#f6d992;text-align:center'><b>Date {}  </b></td><tr>" .format(shift2,ec1,date)
	shift1=1
	for i in shift:
		shift1+=1
	data += "<tr><td rowspan=2 style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center;'>Parent Department</td><td rowspan=2 style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center;'>Department</td><td colspan={} style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>Shift</td><td colspan={} style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>Category</td><td rowspan=2 style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>CheckOut</td></tr>".format(shift1,ec1)        
	data += "<tr>"
	for i in shift:
		data += "<td style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>{}</td>".format(i.name)
	data += "<td style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>Total Present</td>"        
	
	ec=frappe.get_all("Employee Category",{'name':('not in',['Sub Staff','Director'])},['*'])
	for i in ec:
		data += "<td style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>{}</td>".format(i.name)
	data +="</tr>"

	total = 0
	department = frappe.get_all("Department", {'disabled': ('!=', 1),"parent_department":"All Departments"}, ['name'])        
	for d in department:
		length=2
		department1 = frappe.get_all("Department", {'disabled': ('!=', 1),"parent_department":d.name}, ['name'])
		for dep in department1:
			length+=1
		frappe.errprint(length)
		parent_dep=d.name
		total_pre=0
		total_cl=0
		total_trainee=0
		total_ops=0
		total_staff=0
		totl_ch_out=0
		data += "<tr><td rowspan={} style='border: 1px solid black;text-align:left'>{}</td><td style='border: 1px solid black;text-align:center'></td>".format(length,d.name)
		for i in shift:
			shift_attendance_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND shift = %s
				AND department = %s
				AND in_time IS NOT NULL

			""", (date, i.name, d.name), as_dict=True)
			shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
			data += "<td style='border: 1px solid black;text-align:center'>{}</td>".format(shift_attendance)
		staff_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND custom_employee_category IN ("Staff", "Sub Staff", "Director")
			AND department = %s
			AND in_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		staff = staff_count[0].count if staff_count else 0
		ops_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND custom_employee_category IN ("Operator")
			AND department = %s
			AND in_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		ops = ops_count[0].count if ops_count else 0
		aps_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND custom_employee_category IN ("Apprentice")
			AND department = %s
			AND in_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		trainee = aps_count[0].count if aps_count else 0
		cl_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND custom_employee_category IN ("Contractor")
			AND department = %s
			AND in_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		cl = cl_count[0].count if cl_count else 0
		
		checkout_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND department = %s
			AND in_time IS NOT NULL
			AND out_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		ch_out = checkout_count[0].count if checkout_count else 0
		total += (staff+ops+trainee+cl)
		total_pre+=(staff+ops+trainee+cl)
		total_cl+=cl
		total_trainee+=trainee
		total_ops+=ops
		total_staff+=staff
		totl_ch_out+=ch_out
		data += "<td style='border: 1px solid black;text-align:center;background-color:#ADD8E6'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center;background-color:#BACC81'>%s</td>" % ((staff+ops+trainee+cl),cl,trainee,ops,staff,ch_out)
		data += '</tr>'
		department = frappe.get_all("Department", {'disabled': ('!=', 1),"parent_department":d.name}, ['name'])
		for d in department:
			data += "<tr><td style='border: 1px solid black;text-align:center'>%s</td>"%(d.name)
			for i in shift:
				shift_attendance_count = frappe.db.sql("""
					SELECT COUNT(*) AS count
					FROM `tabAttendance`
					WHERE attendance_date = %s
					AND shift = %s
					AND department = %s
					AND in_time IS NOT NULL
				""", (date, i.name, d.name), as_dict=True)
				shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
				data += "<td style='border: 1px solid black;text-align:center'>{}</td>".format(shift_attendance)
			staff_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND custom_employee_category IN ("Staff", "Sub Staff", "Director")
				AND department = %s
				AND in_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			staff = staff_count[0].count if staff_count else 0
			ops_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND custom_employee_category IN ("Operator")
				AND department = %s
				AND in_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			ops = ops_count[0].count if ops_count else 0
			aps_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND custom_employee_category IN ("Apprentice")
				AND department = %s
				AND in_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			trainee = aps_count[0].count if aps_count else 0
			cl_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND custom_employee_category IN ("Contractor")
				AND department = %s
				AND in_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			cl = cl_count[0].count if cl_count else 0
			checkout_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND department = %s
				AND in_time IS NOT NULL
				AND out_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			ch_out = checkout_count[0].count if checkout_count else 0
			total += (staff+ops+trainee+cl)
			total_pre+=(staff+ops+trainee+cl)
			total_cl+=cl
			total_trainee+=trainee
			total_ops+=ops
			total_staff+=staff
			totl_ch_out+=ch_out
			data += "<td style='border: 1px solid black;text-align:center;background-color:#ADD8E6'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center;background-color:#BACC81'>%s</td></tr>" % ((staff+ops+trainee+cl),cl,trainee,ops,staff,ch_out)
		data += "<tr style='border: 1px solid black;text-align:center;background-color:#C0C0C0'><td style='border: 1px solid black;text-align:center'>Total</td>"
		for i in shift:
			shift_count=0
			shift_attendance_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND shift = %s
				AND department = %s
				AND in_time IS NOT NULL

			""", (date, i.name, parent_dep), as_dict=True)
			shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
			shift_count+=shift_attendance
			department = frappe.get_all("Department", {'disabled': ('!=', 1),"parent_department":parent_dep}, ['name'])
			for d in department:
				shift_attendance_count = frappe.db.sql("""
					SELECT COUNT(*) AS count
					FROM `tabAttendance`
					WHERE attendance_date = %s
					AND shift = %s
					AND department = %s
					AND in_time IS NOT NULL

				""", (date, i.name, d.name), as_dict=True)
				shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
				shift_count+=shift_attendance
			data += "<td style='border: 1px solid black;text-align:center'>{}</td>".format(shift_count)
		data+="<td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td></tr>" % (total_pre,total_cl,total_trainee,total_ops,total_staff,totl_ch_out)
	colspan=(shift2)-2
	data += "<tr><td colspan = {} style='border: 1px solid black;text-align:left'>Total Present</td><td colspan=6 style='border: 1px solid black;text-align:left'>{}</td></tr>" .format(colspan,total)
	data += "</table>"
	return data

import frappe
from datetime import date
@frappe.whitelist()
# when already leave application present in draft status for existing balance, then error will be thrown

def restrict_for_zero_balance(doc, method):
	today = date.today()

	
	start_of_year = date(today.year, 1, 1)

	end_of_year = date(today.year, 12, 31)
	if doc.is_new() and doc.leave_type!='Leave Without Pay': 
		total_leave_days_present=0
		total_lbalance=doc.leave_balance
		draft_leave_applications = frappe.get_all("Leave Application", {"employee": doc.employee,"workflow_state":('in',['Draft','Pending For HOD']),"leave_type": doc.leave_type,'from_date':('between',(start_of_year,end_of_year)),'to_date':('between',(start_of_year,end_of_year))},["*"])
		for i in draft_leave_applications:
			total_leave_days_present+=i.total_leave_days
		total_leave_days_present += doc.total_leave_days
		available=total_lbalance-total_leave_days_present
		if available < 0 :
			frappe.throw("Insufficient leave balance for this leave type")

@frappe.whitelist()
def att_request_cancel(doc, method):
	# empty the field attendance request during cancel
	att=frappe.db.get_value("Attendance",{'attendance_request':doc.name},['name'])
	if att:
		attendance = frappe.db.get_value('Attendance', {
			'employee': doc.employee,
			'attendance_date': doc.from_date,
			'docstatus': ("!=", 2)
		}, ['name'])
		frappe.db.set_value('Attendance',att,'attendance_request','')


# restriction to apply leave after a leave day
@frappe.whitelist()
def condition_for_la(doc,method):
	diff = date_diff(today(), doc.from_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager"]]})
	if not role:
		employee_category = frappe.db.get_value("Employee",doc.employee,"employee_category")
		if employee_category in ["Staff","Sub Staff"]:
			if diff > 3:
				frappe.throw("The Leave Application must be apply within 3 days from the leave date")
		if employee_category in ["Apprentice","Operator", "Contractor", "Trainee"]:
			if diff > 0:
				frappe.throw("Leave applications must be applied on or before the same day.")

@frappe.whitelist()
def return_items(doctype,docname):
	doc = frappe.get_doc(doctype,docname)
	return doc.items
@frappe.whitelist()
# restriction to apply attendance request when days exceeded 3 
def condition_for_ar(doc,method):
	diff = date_diff(today(), doc.from_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager","HOD"]]})
	if not role:
		if diff > 3:
			frappe.throw("The Attendance Request must be apply within 3 days")

@frappe.whitelist()
# restriction to apply comp leave request when days exceeded 3 
def condition_for_compoff_lr(doc,method):
	diff = date_diff(today(), doc.work_from_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager","HOD"]]})
	if not role:
		if diff > 3:
			frappe.throw("The Compensatory Leave Request must be apply within 3 days")

@frappe.whitelist()
# restriction to apply attendance permission when days exceeded 3 
def condition_for_ap(doc,method):
	diff = date_diff(today(), doc.permission_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager","HOD"]]})
	if not role:
		if diff > 3:
			frappe.throw("The Attendance Permission must be apply within 3 days")

@frappe.whitelist()
# restriction to apply night shift auditors plan swapping when days exceeded 3 
def condition_for_nsaps(doc,method):
	diff = date_diff(today(), doc.requesting_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager","HOD"]]})
	if not role:
		if diff > 3:
			frappe.throw("The Night Shift Auditors Plan Swapping must be apply within 3 days")

@frappe.whitelist()
# method to return the table in the leave application with the ot value and balance
def get_ot_balance(custom_employee,custom_from_date,custom_to_date):
	data = ''
	OTBalance = frappe.qb.DocType("OT Balance")
	ot_balance = (
		frappe.qb.from_(OTBalance)
		.select(OTBalance.employee, OTBalance.total_ot_hours, OTBalance.comp_off_pending_for_approval,OTBalance.comp_off_used,OTBalance.ot_balance)
		.where(
			(OTBalance.employee == custom_employee)
			& ((custom_from_date >= OTBalance.from_date) & (custom_to_date <= OTBalance.to_date))
		)
	).run(as_dict=True)
	if ot_balance and ot_balance[0]:
		data += '<br><br>'
		data += '<table border=1 width=100%>'
		data += '<tr style="text-align:center;background-color:#ff9248;color:#FFFFFF"><td>Total OT Hours</td><td>C-OFF (Pending for Approval) in Day(s)</td><td>C-OFF Used in Day(s)</td><td>OT Balance Hours</td></tr>'
		data += '<tr style="text-align:center;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(ot_balance[0].total_ot_hours,ot_balance[0].comp_off_pending_for_approval,ot_balance[0].comp_off_used,ot_balance[0].ot_balance)
		data += '</table><br><br>'
	else:
		data += '<p style="text-align:center;"><b>OT balance is not available</b></p>'
	return data

@frappe.whitelist()
#checks the OT balance is sufficient to apply C-OFF from OT
def validate_ot(employee,total_leave_days,from_date,to_date,employee_category):
	result=2
	OTBalance = frappe.qb.DocType("OT Balance")
	ot_balance = (
		frappe.qb.from_(OTBalance)
		.select(OTBalance.ot_balance)
		.where(
			(OTBalance.employee == employee)
			& ((from_date >= OTBalance.from_date) & (to_date <= OTBalance.to_date))
		)
	).run(as_dict=True)
	if ot_balance and ot_balance[0]:
		if float(total_leave_days)*float(8) > float(ot_balance[0].ot_balance):
			result=frappe.throw("Insufficient OT Balance to apply for C-OFF")
			return result
		else:
			if employee_category not in["Staff","Sub Staff"]:
				frappe.errprint(float(total_leave_days)*float(8))
				if frappe.db.exists("Leave Allocation",{'employee':employee,'leave_type':"Compensatory Off",'from_date':['<=', from_date],'to_date':['>=', to_date],'docstatus':("!=",2)}):
					lal=frappe.get_doc("Leave Allocation",{'employee':employee,'leave_type':"Compensatory Off",'from_date':['<=', from_date],'to_date':['>=', to_date],'docstatus':("!=",2)})
					lal.new_leaves_allocated = lal.new_leaves_allocated + float(total_leave_days)
					# frappe.db.set_value('Leave Allocation',lal.name,'new_leaves_allocated',lal.new_leaves_allocated + float(total_leave_days))
					# frappe.db.set_value('Leave Allocation',lal.name,'total_leaves_allocated',lal.total_leaves_allocated + float(total_leave_days))
					lal.save(ignore_permissions=True)
					lal.submit()
				else:
					from_date = datetime.strptime(from_date, "%Y-%m-%d")
					last_date_of_year = date(from_date.year, 12, 31)
					lal=frappe.new_doc("Leave Allocation")
					lal.employee=employee
					lal.leave_type='Compensatory Off'
					lal.from_date=from_date
					lal.to_date=last_date_of_year
					lal.new_leaves_allocated=total_leave_days
					lal.save(ignore_permissions=True)
					lal.submit()

@frappe.whitelist()
#Returns the number of days between the custom from date and to date in Leave Application
def get_number_of_leave_days(
	custom_employee: str,
	custom_from_date: datetime.date,
	custom_to_date: datetime.date,
	custom_half_day: Union[int, str, None] = None,
	custom_half_day_date: Union[datetime.date, str, None] = None,
	holiday_list: Optional[str] = None,
) -> float:
	"""Returns number of leave days between 2 dates after considering half day and holidays
	(Based on the include_holiday setting in Leave Type)"""
	number_of_days = 0
	if cint(custom_half_day) == 1:
		if getdate(custom_from_date) == getdate(custom_to_date):
			number_of_days = 0.5
		elif custom_half_day_date and getdate(custom_from_date) <= getdate(custom_half_day_date) <= getdate(custom_to_date):
			number_of_days = date_diff(custom_to_date, custom_from_date) + 0.5
		else:
			number_of_days = date_diff(custom_to_date, custom_from_date) + 1
	else:
		number_of_days = date_diff(custom_to_date, custom_from_date) + 1

	return number_of_days

@frappe.whitelist()
#returns the Leave Type in the custom feild based on their Leave Balance in Leave Ledger Entry 
def return_select_options(employee):
	from datetime import datetime
	select_option = []
	date=today()
	current_datetime = datetime.now()
	current_year = current_datetime.year
	frappe.errprint(current_year)
	employee_category = frappe.db.get_value('Employee',{'name':employee},'employee_category')
	leave = frappe.db.sql("""
		SELECT leave_type, SUM(leaves) AS total_leaves
		FROM `tabLeave Ledger Entry`
		WHERE docstatus != '2'
		AND employee = %s
		AND from_date <= %s
		AND to_date >= %s
		GROUP BY leave_type
		HAVING total_leaves > 0
		ORDER BY leave_type
	""", (employee, date, date), as_dict=1)

	if employee_category not in ["Staff","Sub Staff"]:
		select_option = ["Comp-off from OT","Leave Without Pay"]
		if leave:
			for l in leave:
				select_option.append(l['leave_type'])
	else:
		select_option = ["Leave Without Pay"]
		if leave:
			for l in leave:
				select_option.append(l['leave_type'])
	return select_option

@frappe.whitelist()
#returns the Leave Type in the custom feild based on their Leave Balance in Leave Ledger Entry 
def get_the_employee_category(employee):
	employee_category = frappe.db.get_value('Employee',{'name':employee},'employee_category')
	user_id = frappe.db.get_value('Employee',{'name':employee},'user_id')
	designation = frappe.db.get_value('Employee',{'name':employee},'designation')
	employee_name = frappe.db.get_value('Employee',{'name':employee},'employee_name')
	department = frappe.db.get_value('Employee',{'name':employee},'department')
	company = frappe.db.get_value('Employee',{'name':employee},'company')
	data =[employee_category,user_id,designation,employee_name,department,company]
	return data

@frappe.whitelist()
#update the draft c-off application count and approved c-off from OT in OT Balance
def otbalance(doc, method):
	month_start = get_first_day(doc.from_date)
	month_end = get_last_day(doc.from_date)
	draft_leave_applications = frappe.get_all(
		"Leave Application",
		filters={
			'employee': doc.employee,
			'from_date': ('between', [month_start, month_end]),
			'to_date': ('between', [month_start, month_end]),
			'workflow_state': 'Pending For HOD',
			'custom_select_leave_type':'Comp-off from OT'
		},
		fields=["total_leave_days"]
	)
	approved_leave_applications = frappe.get_all(
		"Leave Application",
		filters={
			'employee': doc.employee,
			'from_date': ('between', [month_start, month_end]),
			'to_date': ('between', [month_start, month_end]),
			'workflow_state': 'Approved',
			'custom_select_leave_type':'Comp-off from OT'
		},
		fields=["total_leave_days"]
	)
	total_draft_leave_days = sum([i['total_leave_days'] for i in draft_leave_applications])
	total_approved_leave_days = sum([i['total_leave_days'] for i in approved_leave_applications])
	if frappe.db.exists("OT Balance", {'employee': doc.employee, 'from_date': month_start, 'to_date': month_end}):
		otb = frappe.get_doc("OT Balance", {'employee': doc.employee, 'from_date': month_start, 'to_date': month_end})
		otb.comp_off_pending_for_approval = float(total_draft_leave_days)
		otb.comp_off_used = float(total_approved_leave_days)
		otb.ot_balance =float(otb.total_ot_hours)-((float(total_draft_leave_days)*float(8)) + (float(total_approved_leave_days)*float(8)))
		otb.save(ignore_permissions=True)

@frappe.whitelist()
def cancel_leave_application(doc, method):
	if doc.custom_select_leave_type=="Comp-off from OT":
		leave_allocation = frappe.get_doc("Leave Allocation", {
			'employee': doc.custom_employee2,
			'leave_type': "Compensatory Off",
			'from_date': ['<=', doc.from_date],
			'docstatus': ("!=", 2)
		})
		
		if leave_allocation:
			leave_allocation.new_leaves_allocated += float(doc.custom_total_leave_days)
			leave_allocation.save(ignore_permissions=True)
			frappe.db.commit()

		OTBalance = frappe.get_doc("OT Balance", {
			'employee': doc.employee,
			'from_date': ['<=', doc.from_date],
			'to_date': ['>=', doc.to_date]
		})

		if OTBalance:
			OTBalance.ot_balance += float(doc.custom_total_leave_days) * 8
			OTBalance.comp_off_used-=doc.custom_total_leave_days
			OTBalance.save(ignore_permissions=True)
			frappe.db.commit()

@frappe.whitelist()
#send a mail alert to the users in the Custom Settings if the Item is below the Safety Stock
def mail_alert_for_safety_stock():
	item = frappe.get_all("Item",{"disabled":0,"safety_stock":("!=",0)},["name","safety_stock"])
	data = ""
	count=0
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan=3><center>Stock Details</center></th></tr>'
	data += '<tr><td style="padding:1px;border: 1px solid black" colspan =1><b>Item Code</b></td><td style="padding:1px;border: 1px solid black" colspan=1><b>Safety Stock</b></td><td style="padding:1px;border: 1px solid black" colspan=1><b>Available Qty</b></td></tr>'
	for i in item:
		stockqty = frappe.db.sql(""" select item_code,sum(actual_qty) as qty from `tabBin` where item_code = '%s' """%(i.name),as_dict = 1)[0]
		if stockqty['qty']:
			stockqty['qty'] = stockqty['qty']
		else:
			stockqty['qty'] =0
		if i.safety_stock >= stockqty['qty']:
			count+=1
			data += '''  
			<tr><td style="padding:1px;border: 1px solid black" colspan =1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td></tr>'''%(i.name,i.safety_stock,stockqty['qty'] or 0)
	data += '</table>'
	doc = frappe.get_doc("Custom Settings", "Custom Settings")
	emails = []
	cc_mail=[]
	for d in doc.mail_settings:
		if d.report_name == "Stock Details":
			emails.extend(d.recipients.split('\n'))
			cc_mail.extend(d.cc.split('\n'))
	if count > 0:
		frappe.sendmail(
			recipients=emails,
			cc = cc_mail if d.cc else '',
			subject='Stock Details',
			message="""Dear Sir/Mam,<br><br>
				Kindly Check below Item list qty<br>{0}
				""".format(data)
		)
	# user = frappe.db.sql("""
	#     SELECT `tabUser`.name as name
	#     FROM `tabUser`
	#     LEFT JOIN `tabHas Role` ON `tabHas Role`.parent = `tabUser`.name 
	#     WHERE `tabHas Role`.Role = "Stock Manager" AND `tabUser`.enabled = 1
	# """, as_dict=True)
	# if data:
	#     for i in user:
	#         frappe.sendmail(
	#             recipients=["jenisha.p@groupteampro.com"],
	#             subject='Stock Details',
	#             message="""Dear Sir/Mam,<br><br>
	#                 Kindly Check below Item list qty<br>{0}
	#                 """.format(data)
	#         )


@frappe.whitelist()
def return_options():
	from datetime import datetime
	select_option = []
	current_datetime = datetime.now()
	current_year = current_datetime.year
	frappe.errprint(current_year)
	leave = frappe.db.sql("""
		SELECT leave_type, SUM(leaves) AS total_leaves
		FROM `tabLeave Ledger Entry`
		WHERE docstatus != '2'
		AND employee = 'H0002'
		AND YEAR(from_date) <= '2024-01-01'
		AND YEAR(to_date) >= '2024-12-31'
		GROUP BY leave_type
		HAVING total_leaves > 0
		ORDER BY leave_type
	""", as_dict=1)
	employee_category=frappe.db.get_value("Employee",{'name':'H0002'},['employee_category'])
	if employee_category not in ["Staff","Sub Staff"]:
		select_option = ["Comp-off from OT","Leave Without Pay"]
		if leave:
			for l in leave:
				select_option.append(l['leave_type'])
	else:
		select_option = ["Leave Without Pay"]
		if leave:
			for l in leave:
				select_option.append(l['leave_type'])
	return select_option

# @frappe.whitelist()
# def emp_category_update_in_leaveallocation():
# 	leaves=frappe.db.get_all("Leave Allocation",{"leave_type":"Sick Leave","docstatus":"1","from_date":"01-01-2024","to_date":"31-12-2024"},["*"])
# 	ind=0
# 	for i in leaves:
# 		emp=frappe.db.get_value("Employee",{"name":i.employee},["employee_category"])
# 		frappe.db.set_value("Leave Allocation",i.name,"custom_employee_category",emp)
# 		ind+=1

@frappe.whitelist()
#return the last execution time of attendance cron
def update_last_execution():
	doc=frappe.db.get_value("Scheduled Job Log",{"scheduled_job_type":"mark_attendance.mark_att","status":"Complete"},["creation"])
	if doc:
		return doc
	

# @frappe.whitelist()
# #use to create user after saving the employee MIS
# def create_user(employee, first_name, employee_name,date_of_birth,gender):
# # def create_user():
#     # employee = "jjjjj1234"
#     # first_name ="test" 
#     # employee_name ="employee_name"
#     password = "Wonjin@2024"
#     # gender ="Female"
#     # date_of_birth ="2003-04-29"
#     email_str = str(employee) + '@onegeneindia.in'

#     # Check if user already exists
#     existing_user = frappe.get_all('User', filters={'email': email_str})
	
#     if existing_user:
#         return {"message": f"User with email {email_str} already exists. User has been updated."}
#     else:
#         # If user does not exist, create a new user
#         user_doc = frappe.new_doc("User")
#         user_doc.email = email_str
#         user_doc.first_name = first_name
#         user_doc.username = employee
#         user_doc.user_type = "System User"
#         user_doc.birth_date = date_of_birth
#         user_doc.gender = gender
#         user_doc.insert(ignore_permissions=True)
#         # frappe.utils.password.update_password(user_doc.name, password)
#         from frappe.utils.password import update_password
#         update_password(user=user_doc.name, pwd=password)
#         # frappe.db.get_value('')
#         # update_password(user_doc.name, password)
#         # frappe.db.set_value('User',user_doc.name,'new_password',password)
#         # frappe.db.set_value('User',user_doc.name,'password_for_admin',password)
#         # Set create_user_permission for the employee
#     frappe.db.set_value('Employee', {'name': employee}, 'user_id', email_str)
#     frappe.db.set_value('Employee', {'name': employee}, 'create_user_permission', 1)
	
#     # user_id =frappe.db.get_value('User', {'name': email_str}, 'password_for_admin')
#     # print(user_id)
#     return email_str

@frappe.whitelist()
#send a mail alert to PS and API if any scheduled job failed
def schedule_log_fail(doc,method):
	if doc.status=='Failed':
		message = """
		The schedule Job type <b>{}</b> is failed.<br> Kindly check the log <b>{}</b>
		""".format(doc.scheduled_job_type,doc.name)
		frappe.sendmail(
				recipients=["pavithra.s@groupteampro.com","abdulla.pi@groupteampro.com"],
				subject='Scheduled Job type failed(ONEGENE)',
				message=message
			)
		
@frappe.whitelist()
#send a mail alert to PS and API if any scheduled job failed
def update_early():
	frappe.db.set_value('Attendance','HR-ATT-2024-420334','custom_early_out_time','00:00:00')
import datetime
import frappe

@frappe.whitelist()
# def get_absent_count():
def get_absent_count(start_date,end_date,employee):
	import datetime
	# start_date='2025-01-01'
	# end_date='2025-01-31'
	# employee='H0004'
	start_date=str(start_date)
	end_date=str(end_date)
	start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
	end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
	absent_count = 0
	current_week_start = start_date
	while current_week_start <= end_date:
		current_week_end = current_week_start + datetime.timedelta(days=(6 - current_week_start.weekday()))
		if current_week_end>=end_date:
			current_week_end=end_date
		sdate=current_week_start.date()
		edate=current_week_end.date()
		absent_attendance = frappe.db.get_all("Attendance",{'attendance_date': ('between', (sdate, edate)),'status': "Absent",'employee': employee,'docstatus':['!=',2]},['*'])
		woff=0
		for ab in absent_attendance:
			hh=check_holiday(ab.attendance_date,employee)
			if hh is None and ab.leave_application is None and ab.attendance_request is None:
				print(ab.attendance_date)
				woff+=1
		if woff>0:
			absent_count+=1
		current_week_start = current_week_end + datetime.timedelta(days=1)
	return absent_count
from datetime import datetime
import calendar
@frappe.whitelist()
def allocate_el_automatically():
	# frappe.db.set_value("Employee",'4321','employee_category','Staff')
	emp=frappe.db.get_all("Employee",{'Status':'Active'},['name','date_of_joining','employee_category'])
	current_date = datetime.now().date()  # Get the current date
	current_year = current_date.year
	current_month = current_date.month
	# if current_month > 1:
	#     previous_month = current_month-1
	#     start_date_month = datetime(current_year, previous_month, 1).date()
	#     end_date_month = datetime(current_year, previous_month, calendar.monthrange(current_year, previous_month)[1]).date()
	#     start_date = datetime(current_year+1, 1, 1).date()
	#     end_date = datetime(current_year+1, 12, 31).date()
	# else:
	#     previous_month = 12
	#     start_date_month = datetime(current_year - 1, previous_month, 1).date()
	#     end_date_month = datetime(current_year - 1, previous_month, calendar.monthrange(current_year - 1, previous_month)[1]).date()
	#     start_date = datetime(current_year, 1, 1).date()
	#     end_date = datetime(current_year, 12, 31).date()
	
	# for e in emp:
	#     count=0
	#     current_date = datetime.now().date()
	#     doj = e.date_of_joining
	#     diff = current_date - doj
	#     years = diff.days / 365.25 
	#     if int(years)>0:
	#         attendance_filters = {
	#             "attendance_date": ("between", (start_date_month, end_date_month)),
	#             'employee': e['name'],
	#             'status': 'Present',
	#             'docstatus':['!=',2]
	#         }
	#         att = frappe.db.count("Attendance", filters=attendance_filters)
	#         act=32
	#         if e.employee_category in ['Sub Staff','Staff','Operator']:
	#             act=20
	#         elif e.employee_category in ['Apprentice','Trainee']:
	#             act=30
	#         else:
	#             act=32
	#         if att>=act:
	#             count=1
	#         if count>0:
	#             if not frappe.db.exists('Leave Allocation',{'leave_type':'Earned Leave','from_date':start_date,'to_date':end_date,'docstatus':['!=',2],'employee':e.name}):
	#                 allow = frappe.new_doc("Leave Allocation")
	#                 print("EL")
	#                 allow.employee = e.name
	#                 allow.company = e.company
	#                 allow.leave_type = "Earned Leave"
	#                 allow.from_date = start_date
	#                 allow.to_date = end_date
	#                 allow.new_leaves_allocated = count
	#                 allow.total_leaves_allocated = count
	#                 allow.insert()
	#                 allow.save(ignore_permissions=True)
	#                 allow.submit()	
	#                 frappe.db.commit
	#             else:
	#                 allow=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':start_date,'to_date':end_date,'docstatus':['!=',2],'employee':e['name']},['name'])
	#                 leaves=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':start_date,'to_date':end_date,'docstatus':['!=',2],'employee':e['name']},['new_leaves_allocated'])
	#                 tot=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':start_date,'to_date':end_date,'docstatus':['!=',2],'employee':e['name']},['total_leaves_allocated'])
	#                 tot+=count
	#                 leaves+=count
	#                 frappe.db.set_value('Leave Allocation',allow,'new_leaves_allocated',leaves)
	#                 frappe.db.set_value('Leave Allocation',allow,'total_leaves_allocated',leaves)


@frappe.whitelist()
def create_job_fail():
	job = frappe.db.exists('Scheduled Job Type', 'lwf_creation')
	if not job:
		emc = frappe.new_doc("Scheduled Job Type")
		emc.update({
			"method": 'onegene.onegene.custom.create_lwf',
			"frequency": 'Monthly',
		})
		emc.save(ignore_permissions=True)


@frappe.whitelist()
def cron_failed_method():
	cutoff_time = datetime.now() - timedelta(minutes=5)
	failed_jobs = frappe.get_all(
		"Scheduled Job Log",
		filters={
			"status": "Failed",
			"creation": [">=", cutoff_time]
		},
		fields=["scheduled_job_type"]
	)
	unique_job_types = set()
	for job in failed_jobs:
		unique_job_types.add(job['scheduled_job_type'])

	for job_type in unique_job_types:
		frappe.sendmail(
			recipients = ["erp@groupteampro.com","jenisha.p@groupteampro.com","pavithra.s@groupteampro.com","gifty.p@groupteampro.com"],
			subject = 'Failed Cron List - Wonjin',
			message = 'Dear Sir / Mam <br> Kindly find the below failed Scheduled Job  %s'%(job_type)
		)

@frappe.whitelist()
def update_coff():
	count=0
	tot=0
	att= frappe.db.get_all("Attendance",{'attendance_date':'2025-01-12','status':'Present','docstatus':1},['*'])
	for a in att:
		# count+=1
		if frappe.db.exists("Attendance",{'attendance_date':'2025-01-16','status':'Absent','employee': a.employee}):
			count+=1
			print(a.employee)
			# if a.employee not in ['AN4594','S0373']:
			# 		la = frappe.new_doc('Leave Application')
			# 		print(a.employee)
			# 		la.employee = a.employee
			# 		la.leave_type = 'Compensatory Off'
			# 		la.custom_employee2 = a.employee
			# 		la.custom_leave_type = 'Compensatory Off'
			# 		la.from_date = '2025-01-15'
			# 		la.to_date = '2025-01-15'
			# 		la.description = 'Created automatically Via Bulk compensation allocation document'
			# 		la.company = 'WONJIN AUTOPARTS INDIA PVT.LTD.'
			# 		la.status = 'Approved'  
			# 		la.insert(ignore_permissions=True)
			# 		la.submit()
			# 		print(la.name)
			# 		print(a.employee)

	return count

from io import BytesIO
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import frappe
from frappe.utils.background_jobs import enqueue

@frappe.whitelist()
def manpower_cost_download():
	filename = "MANPOWER COST"
	args = frappe.local.form_dict
	frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
	enqueue(build_xlsx_response, queue='long', timeout=36000, event='build_xlsx_response',filename=filename,args=args)
	# build_xlsx_response(filename)

def build_xlsx_response(filename,args):
	xlsx_file = make_xlsx(filename,args)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Reports Dashboard',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	# frappe.db.set_value('Reports Dashboard',None,'manpower_cost_report',attached_file.file_url)
	doc = frappe.get_single("Reports Dashboard")
	doc.reload()
	doc.manpower_cost_report = attached_file.file_url
	doc.save()


def apply_border_to_merged_cells(ws, start_row, start_col, end_row, end_col, border):
	for row in range(start_row, end_row + 1):
		for col in range(start_col, end_col + 1):
			cell = ws.cell(row=row, column=col)
			cell.border = border

def make_xlsx(data,args, sheet_name="MANPOWER COST", wb=None, column_widths=None):
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.active
	ws.title = sheet_name

	from_date = datetime.strptime(args.get('from_date'), '%Y-%m-%d')
	to_date = datetime.strptime(args.get('to_date'), '%Y-%m-%d')
	fromdate = datetime.strptime(args.get('from_date'), '%Y-%m-%d').date()
	from_date_formated = fromdate.strftime('%d-%m-%Y')
	todate = datetime.strptime(args.get('to_date'), '%Y-%m-%d').date()
	to_date_formated = todate.strftime('%d-%m-%Y')
	month_year_format = from_date.strftime('%B %Y')
	# if isinstance(from_date, str):
	#     month_year_format_str = datetime.strptime(from_date, '%Y-%m-%d')
	# else:
	#     month_year_format_str = from_date
	# month_year_format = month_year_format_str.strftime('%B %Y')
	# Header
	header_value = f"Manpower Cost Report ({from_date_formated}  to  {to_date_formated})" 
	# Get the last column index (based on date columns)
	num_days = (to_date - from_date).days + 1
	last_column_index = 6 + 1 * 5 - 1  # 5 columns per date (Attn, Amount, OT, Amount OT, Total)

	# Merge the header cells across all columns
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
	
	thin_border = Border(
		left=Side(style="thin"),
		right=Side(style="thin"),
		top=Side(style="thin"),
		bottom=Side(style="thin"),
	)
	apply_border_to_merged_cells(ws, 1, 1, 1, last_column_index, thin_border)
	first_cell = ws.cell(row=1, column=1)
	first_cell.value = header_value
	first_cell.font = Font(bold=True, size=14, color="0000FF")  # Blue color for header_value
	first_cell.alignment = Alignment(horizontal="center", vertical="center")  # Center align header

	# Table Headers
	headers = ["Sl No", "Department", "Block Name", "Sal / month", "MP Plan"]
	bg_fill = PatternFill(start_color="e2efd9", end_color="e2efd9", fill_type="solid")
	
	for i, header in enumerate(headers, 1):
		ws.merge_cells(start_row=2, start_column=i, end_row=4, end_column=i)
		apply_border_to_merged_cells(ws, 2, i, 4, i, thin_border)
		cell = ws.cell(row=2, column=i)
		cell.value = header
		if header in ["Sl No", "MP Plan"]:
			cell.font = Font(bold=True, color="FF0000")  # Red color for headers
		if header in ["Department", "Block Name"]:
			cell.fill = bg_fill
			cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
	
	department_hierarchy = get_department_hierarchy()
	shift_filter = args.get('shift')  # The shift filter is optional
	
	ws.column_dimensions['A'].width = 5
	ws.column_dimensions['B'].width = 23
	ws.column_dimensions['C'].width = 25
	ws.column_dimensions['D'].width = 12

	row_index = 5
	serial_number = 1
	date_columns = []
	# Dates Row
	start_column = 6  # Starting column for dates

	# for day_offset in range(num_days):
	current_date = from_date + timedelta(days=1)
	date_str = current_date.strftime("%d-%m-%Y")

	# Merge and set date
	# ws.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=start_column + 4)
	# apply_border_to_merged_cells(ws, 2, start_column, 2, start_column + 4, thin_border)
	# date_cell = ws.cell(row=2, column=start_column)
	# date_cell.value = ''
	# date_cell.font = Font(bold=True)
	# date_cell.alignment = Alignment(horizontal="center", vertical="center")
	# date_cell.font = Font(bold=True)

	# Employee Categories Row
	employee_category = args.get('employee_category')
	ws.merge_cells(start_row=2, start_column=start_column, end_row=3, end_column=start_column + 4)
	apply_border_to_merged_cells(ws, 2, start_column, 3, start_column + 4, thin_border)
	cat_cell = ws.cell(row=2, column=start_column, value=employee_category)
	cat_cell.alignment = Alignment(horizontal="center", vertical="center")
	cat_cell.border = thin_border
	cat_cell.font = Font(bold=True)

	# Sub-columns below categories
	col_type = ["Attn", "Amount", "OT", "Amount", "Total"]
	for col_offset, col_name in enumerate(col_type):
		col_cell = ws.cell(row=4, column=start_column + col_offset)
		col_cell.value = col_name
		col_cell.font = Font(bold=True)
		col_cell.alignment = Alignment(horizontal="center", vertical="center")
		col_cell.border = thin_border

	# Store the starting column index for attendance
	# date_columns.append((current_date, start_column))
	# start_column += 5
	
	total_att = 0
	total_att_amt = 0
	total_ot = 0
	total_ot_amt = 0
	grand_total = 0
	data_column = 6
	
	for department, blocks in department_hierarchy.items():
		start_row = row_index
		
		dept_total_att = 0
		dept_total_att_amt = 0
		dept_total_ot = 0
		dept_total_ot_amt = 0
		dept_total = 0
		
		cell = ws.cell(row=row_index, column=2, value=department)
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
		for block in blocks:
			
			s_no = ws.cell(row=row_index, column=1, value=serial_number).border = thin_border
			s_no.font = Font(bold=True)
			child = ws.cell(row=row_index, column=3, value=block).border = thin_border
			child.font = Font(bold=True)
			
			base_result = frappe.db.sql("""
				SELECT SUM(base) AS base 
				FROM `tabSalary Structure Assignment` 
				WHERE custom_employee_category = %s 
				AND department = %s
				AND docstatus = 1
			""", (employee_category, block), as_dict=True)

			if base_result and base_result[0].get("base") is not None:
				base = base_result[0].get("base", 0)
				base_percent = base / 30  # Assuming 30 days in a month for calculation
			else:
				base_percent = 0
				
			round_of_base = round(base_percent)
			
			base_per = ws.cell(row=row_index, column=4, value= round_of_base).border = thin_border
			base_per.font = Font(bold=True)
			start_col =6
		# for date, start_col in date_columns:
			# from_date = datetime.strptime(args.get('from_date'), '%Y-%m-%d')
			# to_date = datetime.strptime(args.get('to_date'), '%Y-%m-%d')
			# current_date = from_date + timedelta(days=1)
			# to_date_str = to_date
			# date_str = current_date.strftime('%Y-%m-%d')
			month_str = current_date.strftime("%b")
			year = current_date.year
			frappe.errprint(date_str)
			date_str = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
			to_date_str = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
			manpower_plan = frappe.db.sql("""
											SELECT SUM(business_plan) as plan
											FROM `tabManpower Plan`
											WHERE department = %s
											AND month = %s
											AND year = %s
											""", (block, month_str, year))
			
			manpower = manpower_plan[0][0] if manpower_plan and manpower_plan[0][0] is not None else 0
			manpower = manpower*num_days
			man = ws.cell(row=row_index, column=5, value= manpower).border = thin_border
			man.font = Font(bold=True)

			# Attendance Calculation
			if shift_filter:
				attendance = frappe.db.sql("""
					SELECT COUNT(*) as count 
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
						AND shift = %s
				""", (date_str,to_date_str, block, employee_category, shift_filter), as_dict=True)
			else:
				attendance = frappe.db.sql("""
					SELECT COUNT(*) as count 
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s 
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
				""", (date_str,to_date_str, block, employee_category), as_dict=True)

			att_value = attendance[0]["count"] if attendance else 0

			# Ensure base_percent is rounded properly
			base_percent = round(base / 30) if base_result and base_result[0].get("base") else 0

			# Calculate amount_value using integer multiplication
			att_amount = att_value * base_percent
			
			att = ws.cell(row=row_index, column=start_col, value= att_value).border = thin_border
			att.font = Font(bold=True)
			
			amount_att = ws.cell(row=row_index, column=start_col+1, value= att_amount).border = thin_border
			amount_att.font = Font(bold=True)
			
			if shift_filter:
				ot_result = frappe.db.sql("""
					SELECT SUM(custom_overtime_hours) as ot
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
						AND shift = %s
				""", (date_str,to_date_str, block, employee_category ,shift_filter), as_dict=True)
			else:
				ot_result = frappe.db.sql("""
					SELECT SUM(custom_overtime_hours) as ot
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s 
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
				""", (date_str,to_date_str, block, employee_category), as_dict=True)

			ot_value = ot_result[0]["ot"] if ot_result and ot_result[0]["ot"] else 0

			ot = ws.cell(row=row_index, column=start_col+2, value= round(ot_value, 2)).border = thin_border
			ot.font = Font(bold=True)
			
			if shift_filter:
				attendance_ot = frappe.db.sql("""
					SELECT SUM(custom_overtime_hours) as ot
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
						AND shift = %s
				""", (date_str,to_date_str, block, employee_category, shift_filter), as_dict=True)
			else:
				attendance_ot = frappe.db.sql("""
					SELECT SUM(custom_overtime_hours) as ot
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
				""", (date_str,to_date_str, block, employee_category), as_dict=True)

			sum_ot = attendance_ot[0]["ot"] if attendance_ot and attendance_ot[0]["ot"] is not None else 0
			ot_amount = sum_ot * base_percent
			
			amount_ot = ws.cell(row=row_index, column=start_col+3, value= ot_amount).border = thin_border
			amount_ot.font = Font(bold=True)
			
			total = att_amount + ot_amount  # Total remains precise with integer arithmetic
			
			tot = ws.cell(row=row_index, column=start_col+4, value= total).border = thin_border
			tot.font = Font(bold=True)
			
			dept_total_att += att_value
			dept_total_att_amt += att_amount
			dept_total_ot += ot_value
			dept_total_ot_amt += ot_amount
			dept_total += total
			
			serial_number += 1
			row_index += 1
			# data_column += 5
			
		total_att += dept_total_att
		total_att_amt += dept_total_att_amt
		total_ot += dept_total_ot
		total_ot_amt += dept_total_ot_amt
		grand_total += dept_total
		
		if blocks:
			ws.merge_cells(start_row=start_row, start_column=2, end_row=row_index - 1, end_column=2)
			
		ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
		cell = ws.cell(row=row_index, column=1, value="SUB - TOTAL")
		cell.font = Font(bold=True, color="FF0000")
		cell.fill = bg_fill
		cell.alignment = Alignment(horizontal="right", vertical="center")
		cell.border = thin_border
		frappe.log_error("test", dept_total_att)
		
		if department != list(department_hierarchy.keys())[-1]:  
			ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
			cell = ws.cell(row=row_index, column=1, value="SUB - TOTAL")
			cell.font = Font(bold=True, color="FF0000")
			cell.fill = bg_fill
			cell.alignment = Alignment(horizontal="right", vertical="center")

			col_index = 6
			while col_index <= last_column_index:
				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total_att)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1

				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total_att_amt)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1

				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total_ot)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1

				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total_ot_amt)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1

				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1
			
			row_index += 1
			
		if department ==  list(department_hierarchy.keys())[-2]:
			green = PatternFill(start_color="385723", end_color="385723", fill_type="solid")
			ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
			cell = ws.cell(row=row_index, column=1, value="VARIABLE SALARY TOTAL")
			cell.font = Font(bold=True, color="FFFFFF")
			cell.fill = green
			cell.alignment = Alignment(horizontal="right", vertical="center")
			
			var_col = 6
			while var_col <= last_column_index:
				cell = ws.cell(row=row_index, column=var_col, value=total_att)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
				cell = ws.cell(row=row_index, column=var_col, value=total_att_amt)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
				cell = ws.cell(row=row_index, column=var_col, value=total_ot)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
				cell = ws.cell(row=row_index, column=var_col, value=total_att_amt)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
				cell = ws.cell(row=row_index, column=var_col, value=grand_total)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
			row_index += 1
			
		if department == list(department_hierarchy.keys())[-1]: 

			ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
			brown = PatternFill(start_color="833c0b", end_color="833c0b", fill_type="solid")

			cell = ws.cell(row=row_index, column=1, value="FIXED SALARY TOTAL")
			cell.font = Font(bold=True, color="FFFFFF")
			cell.fill = brown
			cell.alignment = Alignment(horizontal="center", vertical="center")

			col_idx_last = 6
			while col_idx_last <= last_column_index:
				grand = ws.cell(row=row_index, column=col_idx_last, value=total_att)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

				grand = ws.cell(row=row_index, column=col_idx_last, value=total_att_amt)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

				grand = ws.cell(row=row_index, column=col_idx_last, value=total_ot)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

				grand = ws.cell(row=row_index, column=col_idx_last, value=total_ot_amt)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

				grand = ws.cell(row=row_index, column=col_idx_last, value=grand_total)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

			for col in range(1, last_column_index + 1):
				ws.cell(row=row_index, column=col).fill = brown
				ws.cell(row=row_index, column=col).border = thin_border

			row_index += 1

			white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
			ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
			cell = ws.cell(row=row_index, column=1, value="FIXED SALARY TOTAL")
			cell.font = Font(bold=True, color="000000")
			cell.fill = white
			cell.alignment = Alignment(horizontal="center", vertical="center")

			last_col = 6
			while last_col <= last_column_index:
				last = ws.cell(row=row_index, column=last_col, value=total_att)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last.alignment = Alignment(horizontal="center", vertical="center")
				last_col += 1

				last = ws.cell(row=row_index, column=last_col, value=total_att_amt)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last_col += 1

				last = ws.cell(row=row_index, column=last_col, value=total_ot)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last_col += 1

				last = ws.cell(row=row_index, column=last_col, value=total_ot_amt)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last_col += 1

				last = ws.cell(row=row_index, column=last_col, value=grand_total)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last_col += 1

			for col in range(1, last_column_index + 1):
				ws.cell(row=row_index, column=col).fill = white
				ws.cell(row=row_index, column=col).border = thin_border

			row_index += 1

	ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
	apply_border_to_merged_cells(ws, ws.max_row, 1, ws.max_row, 2, thin_border)

	# Adjust column widths
	if column_widths:
		for col_idx, width in enumerate(column_widths, 1):
			ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

	# Save to BytesIO
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)

	return xlsx_file

def get_department_hierarchy():
	departments = frappe.db.get_all(
		"Department",
		filters={"disabled": 0, "name": ["!=", "All Departments"]},
		fields=["name", "parent_department"],
		order_by="parent_department ASC, name ASC"
	)
	department_hierarchy = {}
	for department in departments:
		parent = department["parent_department"]
		child = department["name"]

		if parent not in department_hierarchy:
			department_hierarchy[parent] = []
		department_hierarchy[parent].append(child)

	return department_hierarchy
from datetime import datetime, date
@frappe.whitelist()
def allocate_el():
	emp=frappe.db.get_all("Employee",{'status':'Active','employee_category':('in',('Staff','Sub Staff','Operator','Apprentice','Trainee'))},['name','date_of_joining','employee_category'])
	current_date = datetime.now().date()
	pcount=0
	if current_date.day == 1 and current_date.month == 1:  
		next_year = current_date.year
		first_date_next_year = datetime(next_year, 1, 1).date()
		last_date_next_year = datetime(next_year, 12, 31).date()
		current_year = current_date.year-1
	else:
		next_year = current_date.year+1
		first_date_next_year = datetime(next_year, 1, 1).date()
		last_date_next_year = datetime(next_year, 12, 31).date()
		current_year = current_date.year
	start_date=datetime(current_year, 1, 1).date()
	yesterday=add_days(current_date,-1)
	# print(emp)
	for e in emp:
		print(e.name)
		count=0
		if e.employee_category in ['Sub Staff','Staff','Operator']:
			act=20
		else:
			act=30
		current_date = datetime.now().date()
		doj = e.date_of_joining
		diff = current_date - doj
		years = diff.days / 365.25 
		if int(years)>0:
			if years < 2 and doj.year==current_year-1:
				doj_date=e.date_of_joining.day
				doj_month=e.date_of_joining.month
				doj=datetime(current_year, doj_month, doj_date).date()
				start_date=doj
			else:
				start_date=datetime(current_year, 1, 1).date()
			print(start_date)
			dates=get_dates(start_date,yesterday)
			for date in dates:
				hh=check_holiday(date,e.name)
				if not hh:
					if frappe.db.exists("Attendance",{'docstatus':['!=',2],'status': 'Present','employee': e.name,"attendance_date":date}):
						count+=1
			print(count)
			if count>0:
				leave=count/act
				if leave>=1:
					if not frappe.db.exists('Leave Allocation',{'leave_type':'Earned Leave','from_date':('between',(first_date_next_year,last_date_next_year)),'to_date':('between',(first_date_next_year,last_date_next_year)),'docstatus':['!=',2],'employee':e.name}):
						allow = frappe.new_doc("Leave Allocation")
						allow.employee = e.name
						allow.company = e.company
						allow.leave_type = "Earned Leave"
						allow.from_date = first_date_next_year
						allow.to_date = last_date_next_year
						allow.new_leaves_allocated = int(leave)
						allow.total_leaves_allocated = int(leave)
						allow.insert()
						allow.save(ignore_permissions=True)
						allow.submit()	
						frappe.db.commit
					else:
						allow=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':first_date_next_year,'to_date':last_date_next_year,'docstatus':['!=',2],'employee':e.name},['name'])
						alloc = frappe.get_doc("Leave Allocation",allow)
						alloc.new_leaves_allocated = int(leave)
						alloc.total_leaves_allocated = int(leave)
						alloc.save(ignore_permissions=True)
						frappe.db.commit

@frappe.whitelist()
def enqueue_el_allocation():
	enqueue(allocate_el, queue='long', timeout=6000)

import requests
import json
import frappe
from frappe import _

@frappe.whitelist()
def update_issue_status_from_teampro():
	issues = frappe.get_all(
		"Issue",
		filters={
			"custom_mail_sent_to_teampro": 1,
			"status": ["not in", ["Resolved", "Closed"]],
		},
		fields=["subject", "priority",'name']
	)

	for i in issues:
		issue_doc=frappe.get_doc("Issue",i.name)
		url = "https://erp.teamproit.com/api/resource/Issue"
		headers = {
			"Content-Type": "application/json",
			"Authorization": "token daa4a43f429c844:3b0d3fbc3c5e4ce"
		}
		params = {
			"filters": json.dumps([
				["subject", "=", i.subject],
				["raised_by", "=", "wonjin_corporate@onegeneindia.in"]
			]),
			"fields": json.dumps(["name","task","custom_issue_status"]),
			"limit_page_length": 1000
		}

		try:
			response = requests.get(url, headers=headers, params=params, verify=False)
			res = response.json()
			for issue_data in res.get("data", []):
				issue_doc.custom_issue_id = issue_data.get("name")
				issue_doc.custom_issue_status = issue_data.get("custom_issue_status")
				task_name = issue_data.get("task")
				if not task_name:
					continue
				task_url = "https://erp.teamproit.com/api/resource/Task"
				task_params = {
					"filters": json.dumps([
						["name", "=", task_name]
					]),
					"fields": json.dumps(["creation", "custom_allocated_to", "status", "description","exp_end_date"]),
				}

				try:
					task_response = requests.get(task_url, headers=headers, params=task_params, verify=False)
					task_data = task_response.json()
					for task in task_data.get("data", []):
						print(task_response.text)
						issue_doc.custom_task_id = task_name
						issue_doc.custom_task_status = task.get("status")
						creation_datetime = task.get("creation")  # Example: '2025-04-09 10:38:26.403313'
						if creation_datetime:
							creation_date = datetime.strptime(creation_datetime, "%Y-%m-%d %H:%M:%S.%f").date()
							issue_doc.custom_task_creation_date = creation_date.strftime("%d-%m-%Y")
						issue_doc.custom_target_date = task.get("exp_end_date")
						issue_doc.custom_task_allocated_to = task.get("custom_allocated_to")
						issue_doc.custom_task_description = task.get("description")
				except Exception:
					frappe.log_error(frappe.get_traceback(), "Teampro Task API Call Failed")
			issue_doc.save(ignore_permissions=True)
		except Exception as e:
			frappe.log_error(frappe.get_traceback(), "Teampro API Call Failed")

@frappe.whitelist()
def create_job_fail1():
	job = frappe.db.exists('Scheduled Job Type', 'update_issue_status_from_teampro')
	if not job:
		emc = frappe.new_doc('Scheduled Job Type')
		emc.update({
			"method": 'onegene.onegene.custom.update_issue_status_from_teampro',
			"frequency": 'Cron',
			"cron_format": '*/15 * * * *'
		})
		emc.save(ignore_permissions=True)

@frappe.whitelist()
def issue_closing_mail(name,subject, message, recipients, sender):
	"""
	Send an email to the specified recipients with the given subject and message.
	"""
	if not frappe.db.exists("Email Account", {"email_id": sender}):
		frappe.throw(_("Sender email not configured in Email Account"))

	# Add additional content to the message body
	message_body = f"""
		<div style="font-family: 'Times New Roman', Times, serif; font-size: 14px;">
			<p>Dear Employee,</p>

			<p>A Ticket : {name} with the below mentioned Subject and Description has been resolved</p>

			<p><strong>Subject:</strong> {subject}</p>
			<p><strong>Description:</strong>{message}</p>


			<p>Thanks & Regards,<br>
			Wonjin Team</p>
		</div>
	"""

	try:
		frappe.sendmail(
			recipients=recipients,
			sender=sender,
			subject=subject,
			message=message_body,
			now=True
		)
		return "Email sent successfully"
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), 'Email Sending Failed')
		return "Failed to send email"
	
@frappe.whitelist()
def create_purchase_open_order(doc, method):
	if doc.custom_order_type == "Open":
		new_doc = frappe.new_doc('Purchase Open Order')
		new_doc.purchase_order = doc.name
		new_doc.set('open_order_table', [])
		for po in doc.items:
			new_doc.append("open_order_table", {
				"item_code": po.item_code,
				"delivery_date": po.schedule_date,
				"item_name": po.item_name,
				"qty": po.qty,
				"rate": po.rate,
				"warehouse": po.warehouse,
				"amount": po.amount,
			})
		new_doc.save(ignore_permissions=True)

@frappe.whitelist()
def create_purchase_order_schedule_from_po(doc,method):
	if doc.custom_order_type == "Fixed" and not doc.custom_schedule_table:
		frappe.throw("Schedule not Created")
	if doc.custom_order_type == "Fixed" and doc.custom_schedule_table:
		for schedule in doc.custom_schedule_table:
			if frappe.db.exists('Purchase Order Schedule', {'purchase_order_number': doc.name, 'item_code': schedule.item_code, 'schedule_date': schedule.schedule_date}):
				frappe.throw("Schedule already exists for this item code and schedule date")
			new_doc = frappe.new_doc('Purchase Order Schedule') 
			new_doc.supplier_code = doc.supplier_code
			new_doc.purchase_order_number = doc.name
			new_doc.item_code = schedule.item_code
			new_doc.schedule_date = schedule.schedule_date
			new_doc.qty = schedule.schedule_qty
			for item in doc.items:
				if item.item_code == schedule.item_code:
					new_doc.child_name = schedule.name
					new_doc.schedule_amount = schedule.schedule_qty * item.rate
					new_doc.order_rate = item.rate
					new_doc.pending_qty = schedule.schedule_qty
					new_doc.pending_amount = schedule.schedule_qty * item.rate
			new_doc.save(ignore_permissions=True) 
			new_doc.submit()

	if doc.custom_order_type == "Open" and doc.custom_schedule_table:
		for schedule in doc.custom_schedule_table:
			date_obj = datetime.strptime(schedule.schedule_date, '%Y-%m-%d')
			month = date_obj.strftime('%b') 
			new_doc = frappe.new_doc('Purchase Order Schedule') 
			new_doc.supplier_code = doc.supplier_code
			new_doc.purchase_order_number = doc.name
			new_doc.item_code = schedule.item_code
			new_doc.schedule_date = schedule.schedule_date
			new_doc.schedule_month = month.upper()
			new_doc.qty = schedule.schedule_qty
			for item in doc.items:
				if item.item_code == schedule.item_code:
					new_doc.child_name = schedule.name
					new_doc.schedule_amount = schedule.schedule_qty * item.rate
					new_doc.order_rate = item.rate
					new_doc.pending_qty = schedule.schedule_qty
					new_doc.pending_amount = schedule.schedule_qty * item.rate
			new_doc.save(ignore_permissions=True) 
			new_doc.submit()

@frappe.whitelist()
def	reload_po(doc,method):
	doc.reload()

@frappe.whitelist()
def delete_purchase_order_schedule(doc, method):
	if doc.order_schedule:
		if frappe.db.exists('Purchase Order Schedule', doc.order_schedule):
			schedule = frappe.get_doc('Purchase Order Schedule', doc.order_schedule)
			if schedule.order_type == "Open":
				schedule.cancel()
				doc.save(ignore_permissions=True)

@frappe.whitelist()
def get_open_order_hooks(doc,method):
	if doc.customer_order_type == "Open":
		new_doc = frappe.new_doc('Open Order')
		new_doc.sales_order_number = doc.name
		new_doc.set('open_order_table', [])
		for so in doc.items:
			new_doc.append("open_order_table", {
				"item_code": so.item_code,
				"delivery_date": so.delivery_date,
				"item_name": so.item_name,
				"qty": so.qty,
				"rate": so.rate,
				"warehouse": so.warehouse,
				"amount": so.amount,
			})
		new_doc.save(ignore_permissions=True)

@frappe.whitelist()
def get_previous_purchase_rate(item_code):
	pr = frappe.db.sql("""
		SELECT 
			lci.rate
		FROM 
			`tabPurchase Order Item` lci
		LEFT JOIN 
			`tabPurchase Order` lcv
		ON 
			lci.parent = lcv.name
		JOIN 
			(SELECT 
				item_code, MAX(creation) as latest_creation
			FROM 
				`tabPurchase Order Item`
			WHERE 
				rate IS NOT NULL
			GROUP BY 
				item_code) latest_lci
		ON 
			lci.item_code= latest_lci.item_code
			AND lci.creation = latest_lci.latest_creation
		WHERE 
			lci.item_code = %s and lcv.docstatus !=2
	""", item_code, as_dict=True)

	if pr:
		return pr[0].get('rate')
	else:
		return '0.0'

@frappe.whitelist()
def set_lead_time(item,supplier):
	item_code = frappe.get_doc("Item",item)
	for i in item_code.supplier_items:
		if supplier == i.supplier:
			return i.custom_lead_time_in_days


@frappe.whitelist()
def create_order_schedule_from_po_for_open(item_code, schedule_date, schedule_qty, supplier_code, name, rate, month):
	rate = float(rate)
	schedule_qty = float(schedule_qty)
	new_doc = frappe.new_doc('Purchase Order Schedule') 
	new_doc.supplier_code = supplier_code
	new_doc.order_type = "Open"
	new_doc.purchase_order_number = name
	new_doc.item_code = item_code
	new_doc.schedule_date = schedule_date
	new_doc.qty = schedule_qty
	new_doc.schedule_amount = schedule_qty * rate
	new_doc.order_rate = rate
	new_doc.schedule_month = month.upper()
	new_doc.pending_qty = schedule_qty
	new_doc.pending_amount = schedule_qty * rate
	new_doc.save(ignore_permissions=True) 
	new_doc.submit()
	return("ok")

@frappe.whitelist()
def update_address_in_supplier():
	address_data = frappe.db.get_all('Address',{'address_type':"Billing"},['name'])
	count =0
	for s in address_data:
		address = frappe.get_value("Dynamic Link", {'parent': s.get('name'), 'link_doctype': 'Supplier'}, 'link_name')
		# print(address)
		if address:
			doc = frappe.get_doc('Address',s.get('name'))
			address_display = get_address_display(doc.as_dict())
			# frappe.db.set_value('Customer',{'customer_primary_address':doc.address_title,'customer_primary_address':['is','not set']},'customer_primary_address',doc.name)
			# frappe.db.set_value('Customer',{'customer_primary_address':doc.address_title,'customer_primary_address':['is','not set']},'primary_address',address_display)
			frappe.db.set_value('Supplier',{'name':doc.address_title,'supplier_primary_address':['is','not set']},'supplier_primary_address',doc.name)
			frappe.db.set_value('Supplier',{'name':doc.address_title,'supplier_primary_address':['is','not set']},'primary_address',address_display)        
			# break
	# doc =frappe.get_doc('Address','M.S.N. ENGINEERING ENTERPRISES-Billing')
	# address_display = get_address_display(doc.as_dict())
	# frappe.db.set_value('Supplier','M.S.N. ENGINEERING ENTERPRISES','supplier_primary_address','M.S.N. ENGINEERING ENTERPRISES-Billing')
	# frappe.db.set_value('Supplier','M.S.N. ENGINEERING ENTERPRISES','primary_address',address_display)

@frappe.whitelist()
def get_last_date_of_fiscal():
	last_date = frappe.db.get_value("Fiscal Year", "2025-2026", "year_end_date")
	return last_date

@frappe.whitelist()
def return_conversion(currency,price_list_currency):
	conv_rate = get_exchange_rate(currency, price_list_currency)
	return conv_rate

@frappe.whitelist()
def set_naming_series_po(name, new_name):
    if name != new_name:
        frappe.rename_doc("Purchase Order", name, new_name, force=True)
        frappe.db.commit()
        return new_name

import math
@frappe.whitelist()
def set_box_weight(item,box,item_qty):
	box_qty = frappe.db.get_value("Box Table",{'parent':item,'box':box},['qty'])
	no_of_boxes = 0
	if box_qty:
		no_of_boxes = math.ceil(float(item_qty) / float(box_qty))
	box_weight = frappe.db.get_value("Box",{'name':box},['weight'])
	total_weight_of_boxes = no_of_boxes * box_weight
	total_box_length = frappe.db.get_value("Box",{'name':box},['length']) * no_of_boxes
	total_box_breadth = frappe.db.get_value("Box",{'name':box},['breadth']) * no_of_boxes
	total_box_height = frappe.db.get_value("Box",{'name':box},['height']) * no_of_boxes
	return no_of_boxes, total_weight_of_boxes, total_box_length, total_box_breadth, total_box_height

@frappe.whitelist()
def set_pallet_weight(item,box,pallet,box_qty):
	pallet_qty = frappe.db.get_value("Pallet Table",{'parent':item,'box':box,'pallet':pallet},['qty'])
	no_of_pallets = 0
	if pallet_qty:
		no_of_pallets = math.ceil(float(box_qty) / float(pallet_qty))
	pallet_weight = frappe.db.get_value("Pallet",{'name':pallet},['weight'])
	total_weight_of_pallets = no_of_pallets * pallet_weight
	total_pallet_length = frappe.db.get_value("Pallet",{'name':pallet},['length']) * no_of_pallets
	total_pallet_breadth = frappe.db.get_value("Pallet",{'name':pallet},['breadth']) * no_of_pallets
	total_pallet_height = frappe.db.get_value("Pallet",{'name':pallet},['height']) * no_of_pallets
	return no_of_pallets, total_weight_of_pallets, total_pallet_length, total_pallet_breadth, total_pallet_height