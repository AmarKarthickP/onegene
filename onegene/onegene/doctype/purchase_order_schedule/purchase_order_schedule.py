# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime


from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import frappe
from frappe.query_builder import DocType
import requests
from datetime import date
from time import strptime
import erpnext
import json
from frappe.utils import now
from typing import Dict, Optional, Tuple, Union
from frappe import throw,_
from frappe.utils import flt
from frappe.utils import (
	add_days,
	ceil,
	cint,
	comma_and,
	flt,
	get_link_to_form,
	getdate,
	now_datetime,
	datetime,get_first_day,get_last_day,
	nowdate,
	today,
)
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days,date_diff
from datetime import date, datetime, timedelta
from erpnext.stock.get_item_details import (
	_get_item_tax_template,
	get_conversion_factor,
	get_item_details,
	get_item_tax_map,
	get_item_warehouse,
)
from frappe.model.workflow import get_workflow_name, is_transition_condition_satisfied

class PurchaseOrderSchedule(Document):
	def before_insert(self):
		if self.purchase_order_number and self.item_code:
			order_rate = frappe.db.get_value(
				"Purchase Order Item", 
				{"parent": self.purchase_order_number, "item_code": self.item_code}, 
				"rate"
			)
			self.order_rate = order_rate or 0
		else:
			self.order_rate = 0

		if self.schedule_date:
			if isinstance(self.schedule_date, str):
				try:
					self.schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d %H:%M:%S")
				except ValueError:
					self.schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d")
		
			self.schedule_year = self.schedule_date.year


	def after_insert(self):
		if self.amended_from:
			if frappe.db.exists("Supplier-DN Schedule Trace", {"purchase_order_schedule", self.amended_from}):
				frappe.db.sql(
					"UPDATE `tabSupplier-DN Schedule Trace` SET purchase_order_schedule = %s WHERE purchase_order_schedule = %s",
					(self.name, self.amended_from),
					as_dict=False
				)
				frappe.db.commit()

	def validate(self):
		if self.order_type == "Open":
			current_year = datetime.now().year
			schedule_month = self.schedule_month.upper() if self.schedule_month else ""
			self.schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")
			self.schedule_year = current_year
		if self.order_type == "Fixed":
			if self.schedule_date:
				if isinstance(self.schedule_date, str):
					schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d")
				else:
					schedule_date = self.schedule_date
				self.schedule_year = schedule_date.year
				self.schedule_month = schedule_date.strftime("%b").upper()

		self.pending_qty = self.qty - self.received_qty
		self.schedule_amount = self.order_rate * self.qty
		self.received_amount = self.order_rate * self.received_qty
		self.pending_amount = self.order_rate * (self.qty - self.received_qty)
		if self.qty < self.received_qty:
			frappe.throw(
				"Cannot set Schedule Quantity less than Received Quantity",
			)
		
		# Function for amendment
		if self.amended_from:
			if not self.revised:
				if frappe.db.exists("Purchase Order Schedule", self.amended_from):
					old_doc = frappe.get_doc("Purchase Order Schedule", self.amended_from)
					new_row = self.append('revision', {})
					new_row.revised_on = now_datetime()
					new_row.schedule_qty = old_doc.qty
					new_row.revised_schedule_qty = self.qty
			else:
				if self.docstatus == 1:
					return
				child_doc = self.revision[-1]
				child_doc.revised_schedule_qty = self.qty

	def on_cancel(self):
		if self.qty < self.received_qty:
			frappe.throw(
				"The scheduled quantity has been fully received. Please create a new Order Schedule to make changes.",
				title="Revision Not Allowed"
			)
		if self.order_type == "Open":
			if frappe.db.exists("Purchase Order",self.purchase_order_number):
				po = frappe.get_doc("Purchase Order",self.purchase_order_number)
				po.custom_schedule_table = [
					row for row in po.custom_schedule_table if row.order_schedule != self.name
				]
				po.save(ignore_permissions=True)
			if frappe.db.exists("Purchase Open Order", {"purchase_order": self.purchase_order_number}):
				poo = frappe.get_doc("Purchase Open Order", {"purchase_order": self.purchase_order_number})
				for row in poo.open_order_table:
					if row.item_code == self.item_code:
						if row.qty - self.qty > 0:
							row.qty = row.qty - self.qty
						else:
							row.qty = 1
				poo.save(ignore_permissions=True)

	def on_submit(self):
		if not self.purchase_order_number:
			return

		order_type = frappe.db.get_value("Purchase Order", self.purchase_order_number, 'custom_order_type')
		po = frappe.get_doc("Purchase Order", self.purchase_order_number)

		if self.order_type == "Open":
			# Update or Create Schedule Item
			filters = {
				"item_code": self.item_code,
				"schedule_qty": self.qty,
				"parent": self.purchase_order_number,
				"order_schedule": ("is", None)
			}
			if frappe.db.exists("Purchase Order Schedule Item", filters):
				frappe.db.set_value("Purchase Order Schedule Item", filters, "order_schedule", self.name)
			else:
				frappe.get_doc({
					"doctype": "Purchase Order Schedule Item",
					"parent": self.purchase_order_number,
					"parenttype": "Purchase Order",
					"parentfield": "custom_schedule_table",
					"item_code": self.item_code,
					"schedule_date": self.schedule_date,
					"schedule_month": self.schedule_month,
					"schedule_qty": self.qty,
					"received_qty": self.received_qty,
					"pending_qty": self.pending_qty,
					"order_schedule": self.name
				}).insert(ignore_permissions=True)

		# Update PO Item Qty and Amount if Draft
		if order_type == "Open" and frappe.db.exists("Purchase Order", {'name': self.purchase_order_number, 'docstatus': 0}):
			s_qty = sum(d.qty for d in frappe.get_all("Purchase Order Schedule", {
				"purchase_order_number": self.purchase_order_number,
				"docstatus": 1,
				"supplier_code": self.supplier_code,
				"item_code": self.item_code
			}, ["qty"]))

			# Update PO Item
			frappe.db.set_value("Purchase Order Item", {"parent": self.purchase_order_number, "item_code": self.item_code}, "qty", s_qty)

			# Update PO Doc
			for row in po.items:
				if row.item_code == self.item_code:
					row.qty = s_qty
					row.amount = s_qty * row.rate
			po.save(ignore_permissions=True)

		# Update Open Order Table if PO is Submitted
		if order_type == "Open" and frappe.db.exists("Purchase Order", {'name': self.purchase_order_number, 'docstatus': 1}):
			open_order = frappe.get_doc("Purchase Open Order", {"purchase_order": self.purchase_order_number})
			item_qty = sum(s.qty for s in frappe.get_all("Purchase Order Schedule", {
				"purchase_order_number": self.purchase_order_number,
				"item_code": self.item_code,
				"docstatus": 1
			}, ["qty"]))
			matching_row = next((row for row in open_order.open_order_table if row.item_code == self.item_code), None)
			if matching_row:
				if self.disable_update_items == 1:
					open_order.disable_update_items = 1
				else:
					open_order.disable_update_items = 0
				matching_row.qty = item_qty
				matching_row.amount = item_qty * float(self.order_rate)
				open_order.save(ignore_permissions=True)

		# For Fixed Orders - Validate qty
		if order_type == "Fixed":
			s_qty = sum(d.qty for d in frappe.get_all("Purchase Order Schedule", {
				"purchase_order_number": self.purchase_order_number,
				"supplier_code": self.supplier_code,
				"item_code": self.item_code,
				"docstatus": 1
			}, ["qty"]))

			po_items = frappe.get_all("Purchase Order Item", {
				"parent": self.purchase_order_number,
				"item_code": self.item_code
			}, ["qty", "item_code", "idx"])

			if not po_items:
				frappe.throw(f"Item <b>{self.item_code}</b> not found in the Purchase Order <b>{self.purchase_order_number}</b>")

			po_qty = po_items[0]["qty"]
			if s_qty > po_qty:
				frappe.throw(f"Validation failed: Quantity <b>{s_qty}</b> exceeds Purchase Order quantity <b>{po_qty}</b> for item <b>{self.item_code}</b>.")

	def on_update_after_submit(self):
		self.schedule_amount = self.order_rate * self.qty
		self.received_amount = self.order_rate * self.received_qty
		self.pending_qty = self.qty - self.received_qty
		self.pending_amount = self.order_rate * self.pending_qty

		if frappe.db.exists("Purchase Order", self.purchase_order_number):
			po = frappe.get_doc("Purchase Order", self.purchase_order_number)
			for row in po.custom_schedule_table:
				if row.order_schedule == self.name:
					row.schedule_qty = self.qty
					row.received_qty = self.received_qty
					row.pending_qty = self.pending_qty
			po.save(ignore_permissions=True)



@frappe.whitelist()
def return_items(doctype,docname):
	doc = frappe.get_doc(doctype,docname)
	return doc.items

@frappe.whitelist()
def schedule_list(purchase, item):
	if purchase and item:
		documents = frappe.get_all('Purchase Order Schedule', {'purchase_order_number': purchase, 'item_code': item, "docstatus": 1},
									['schedule_date', 'tentative_plan_1', 'tentative_plan_2', 'qty', 'delivered_qty',
									'pending_qty', 'remarks', 'order_rate'])

		documents = sorted(documents, key=lambda x: x['schedule_date'])
		data = '<table border="1" style="width: 100%;">'
		data += '<tr style="background-color:#D9E2ED;">'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Month</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Date</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Tentative Plan - I</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Tentative Plan - II</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Delivered Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Pending Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Remarks</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Cost Price</b></td>'
		data += '</tr>'
		for doc in documents:
			month_string = doc['schedule_date'].strftime('%B')
			data += '<tr>'
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(month_string)
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['schedule_date'].strftime('%d-%m-%Y'))
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['tentative_plan_1'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['tentative_plan_2'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['qty'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['delivered_qty'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['pending_qty'])
			if doc['remarks']:
				data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['remarks'])
			else:
				data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format('-')
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['order_rate'])
			data += '</tr>'
		data += '</table>'
		return data

@frappe.whitelist()
def get_qty_rate_so(item,purchase):
	so = frappe.db.get_value("Purchase Order Item",{"Parent":purchase,"item_code":item},["rate"])
	return so

@frappe.whitelist()
def get_qty_rate_test():
	s_qty = 0
	po_exist = frappe.db.exists('Purchase Order Schedule',{"purchase_order_number":'PUR-ORD-2025-00012',"supplier_code":'Amar',"item_code":'RDAE1T6SPRWON243'})
	if po_exist:
		exist_po = frappe.get_all("Purchase Order Schedule",{"purchase_order_number":'PUR-ORD-2025-00012',"supplier_code":'Amar',"item_code":'RDAE1T6SPRWON243'},["*"])
		for i in exist_po:
			old_qty = i.qty
			s_qty += old_qty
		print(s_qty)
		frappe.db.set_value("Purchase Order Item",{"parent":'PUR-ORD-2025-00012',"item_code":'RDAE1T6SPRWON243'},"qty",s_qty)			

@frappe.whitelist()
def update_qty_in_open_order(qty, purchase_order, item_code):
	qty = float(qty)
	if frappe.db.exists("Purchase Open Order", {"purchase_order": purchase_order}):
		poo = frappe.get_doc("Purchase Open Order", {"purchase_order": purchase_order})
		for row in poo.open_order_table:
			if row.item_code == item_code:
				row.qty = row.qty + qty
		poo.save(ignore_permissions=True)
		return "ok"

@frappe.whitelist()
def revise_schedule_qty(name, revised_qty, remarks):
	revised_qty = flt(revised_qty)
	doc = frappe.get_doc("Purchase Order Schedule", name)

	if revised_qty < flt(doc.received_qty):
		frappe.throw("Cannot set Schedule Quantity less than Received Quantity")

	doc.append("revision", {
		"revised_on": frappe.utils.now_datetime(),
		"remarks": remarks,
		"schedule_qty": doc.qty,
		"revised_schedule_qty": revised_qty,
		"revised_by": frappe.session.user
	})
	doc.qty = revised_qty
	doc.disable_update_items = 0
	doc.pending_qty = flt(revised_qty) - flt(doc.received_qty)
	doc.schedule_amount = flt(revised_qty) * flt(doc.order_rate)
	doc.received_amount = flt(doc.received_qty) * flt(doc.order_rate)
	doc.pending_amount = (flt(revised_qty) - flt(doc.received_qty)) * flt(doc.order_rate)
	doc.save(ignore_permissions=True)
	# frappe.db.set_value("Purchase Order Schedule", name, "pending_qty", flt(revised_qty) - flt(doc.received_qty))

	self = frappe.get_doc("Purchase Order Schedule", name)
	if self.order_type == "Open" and frappe.db.exists("Purchase Order", {'name': self.purchase_order_number, 'docstatus': 1}):
			open_order = frappe.get_doc("Purchase Open Order", {"purchase_order": self.purchase_order_number})
			item_qty = sum(s.qty for s in frappe.get_all("Purchase Order Schedule", {
				"purchase_order_number": self.purchase_order_number,
				"item_code": self.item_code,
				"docstatus": 1
			}, ["qty"]))
			matching_row = next((row for row in open_order.open_order_table if row.item_code == self.item_code), None)
			if matching_row:
				if self.disable_update_items == 1:
					open_order.disable_update_items = 1
				else:
					open_order.disable_update_items = 0
				matching_row.qty = item_qty
				matching_row.amount = item_qty * float(self.order_rate)
				open_order.save(ignore_permissions=True)

@frappe.whitelist()
def get_query_for_item_table(doctype, txt, searchfield, start, page_len, filters):
	po = filters.get("purchase_order")
	if not po:
		return []

	return frappe.db.sql("""
		SELECT poi.item_code, i.item_name
		FROM `tabPurchase Order Item` poi
		JOIN `tabItem` i ON poi.item_code = i.name
		WHERE poi.parent = %s AND (poi.item_code LIKE %s OR i.item_name LIKE %s)
		LIMIT %s OFFSET %s
	""", (po, f"%{txt}%", f"%{txt}%", page_len, start))
