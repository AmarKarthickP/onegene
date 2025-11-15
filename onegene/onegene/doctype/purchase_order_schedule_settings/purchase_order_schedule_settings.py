# Copyright (c) 2025, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import json
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class PurchaseOrderScheduleSettings(Document):
	pass

import frappe
import uuid

import uuid
from datetime import datetime
import frappe
from frappe.utils.file_manager import get_file
from frappe.exceptions import PermissionError
from frappe import _
from onegene.onegene.doctype.purchase_order_schedule.purchase_order_schedule import revise_schedule_qty

@frappe.whitelist()
def enqueue_upload(file):
	import uuid

	job_id = f"upload_job_{uuid.uuid4().hex}"

	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])
	pps = [pp for pp in pps if pp and pp[0] != "Supplier Code"]
	records=pps
	if len(pps) < 20:
		precheck_purchase_order_upload_without_enqueue(records,file[0],pps,job_id)
		frappe.db.set_single_value('Purchase Order Schedule Settings', 'attach', None)
		return _process_upload(file=file[0], records=pps)
	elif len(pps) <= 500:
		precheck_purchase_order_upload(records,file[0],pps,job_id)
		frappe.db.set_single_value('Purchase Order Schedule Settings', 'attach', None)
		frappe.msgprint(_("Upload is being processed in background..."), alert=True)
		
	else:
		frappe.throw(_("Upload supports only up to 500 rows"), title=_("Too Many Rows"))
@frappe.whitelist()
def precheck_purchase_order_upload_without_enqueue(records,file_id,pps,job_id):
	from datetime import datetime
	error_log = []

	distinct_po_numbers = list({r[1] for r in records if r[1]})
	for idx, pp in enumerate(records, start=1):
		try:
			sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], pp[3]
			s_qty = pp[4]
			
			if not po_number:
				error_log.append(f"Row {idx}: Missing Purchase Order Number.")
				continue

			if not frappe.db.exists("Purchase Order", po_number):
				error_log.append(f"Row {idx}: Purchase Order {po_number} does not exist.")
				continue

			purchase = frappe.get_doc("Purchase Order", po_number)
			if purchase.custom_order_type != "Open":
				error_log.append(f"Row {idx}: PO {po_number} is not of type 'Open'.")
				continue

			schedule_month = schedule_month.upper() if schedule_month else ""
			current_year = datetime.now().year

			try:
				_ = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")
			except Exception:
				error_log.append(f"Row {idx}: Invalid schedule month '{schedule_month}'. Expected format like 'JAN', 'FEB'.")
				continue

		except Exception as e:
			error_log.append(f"Row {idx} failed during validation: {e}")

	if error_log:
		frappe.log_error("\n".join(error_log), "Purchase Order Upload Validation Errors")
		frappe.throw(f'{error_log}')
		# frappe.throw("Validation failed. Upload aborted:\n\n" + "\n".join(error_log))
	

@frappe.whitelist()
def precheck_purchase_order_upload(records,file_id,pps,job_id):
	from datetime import datetime
	error_log = []

	distinct_po_numbers = list({r[1] for r in records if r[1]})
	for idx, pp in enumerate(records, start=1):
		try:
			sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], pp[3]
			s_qty = pp[4]

			if not po_number:
				error_log.append(f"Row {idx}: Missing Purchase Order Number.")
				continue

			if not frappe.db.exists("Purchase Order", po_number):
				error_log.append(f"Row {idx}: Purchase Order {po_number} does not exist.")
				continue

			purchase = frappe.get_doc("Purchase Order", po_number)
			if purchase.custom_order_type != "Open":
				error_log.append(f"Row {idx}: PO {po_number} is not of type 'Open'.")
				continue

			schedule_month = schedule_month.upper() if schedule_month else ""
			current_year = datetime.now().year

			try:
				_ = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")
			except Exception:
				error_log.append(f"Row {idx}: Invalid schedule month '{schedule_month}'. Expected format like 'JAN', 'FEB'.")
				continue

		except Exception as e:
			error_log.append(f"Row {idx} failed during validation: {e}")

	if error_log:
		frappe.log_error("\n".join(error_log), "Purchase Order Upload Validation Errors")
		frappe.throw('')
		# frappe.throw("Validation failed. Upload aborted:\n\n" + "\n".join(error_log))
	else:
		frappe.enqueue(
			method="onegene.onegene.doctype.purchase_order_schedule_settings.purchase_order_schedule_settings._process_upload",
			queue="long",
			timeout=1800,
			is_async=True,
			file=file_id,
			records=pps,
			job_name=job_id,
		)

def _process_upload(file, records):
	from collections import defaultdict
	from datetime import datetime
	import frappe

	docs_to_submit = []
	grouped_docs = defaultdict(list)  
	distinct_po_numbers = []
	total_records = len(records)
	error_log = []
	
	new_count = 0
	update_count = 0
	skipped_count = 0

	for idx, pp in enumerate(records, start=1):
		try:
			sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], pp[3]
			s_qty = pp[4]

			if not po_number:
				continue

			purchase = frappe.get_doc("Purchase Order", po_number)

			if purchase.custom_order_type == "Open":
				current_year = datetime.now().year
				schedule_month = schedule_month.upper() if schedule_month else ""
				schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")
				po_type = frappe.db.get_value("Purchase Order",{"name":po_number},"custom_is_jobcard__subcontracted")
    
				if not frappe.db.exists("Purchase Order Schedule",{"item_code":item,"purchase_order_number":po_number,"schedule_month":schedule_month, "docstatus": 1}):
					doc = frappe.new_doc('Purchase Order Schedule')
					doc.supplier_code = sup_code
					doc.supplier_name = frappe.db.get_value("Supplier", {"supplier_code": sup_code}, "name")
					doc.purchase_order_number = po_number
					doc.po_type = "Job Order" if po_type else "Purchase Order"
					doc.item_code = item
					doc.schedule_date = schedule_date
					doc.qty = s_qty
					doc.schedule_month = schedule_month
					doc.pending_qty = s_qty
					doc.disable_update_items = 1
					doc.save(ignore_permissions=True)

					docs_to_submit.append(doc.name)
					distinct_po_numbers.append(po_number)
					grouped_docs[po_number].append(doc.name)
					new_count += 1
				else:
					if not frappe.db.exists("Purchase Order Schedule",{"item_code":item,"purchase_order_number":po_number,"schedule_month":schedule_month,"qty":s_qty, "docstatus": 1}):
						existing_pos = frappe.db.get_value("Purchase Order Schedule",{"item_code":item,"purchase_order_number":po_number,"schedule_month":schedule_month},['name'])
						revise_schedule_qty(existing_pos,s_qty,'Revised from Purchase Order Schedule Settings')
						distinct_po_numbers.append(po_number)
						update_count += 1
					else:
						skipped_count += 1
			frappe.publish_realtime(
				event='purchase_order_upload_progress',
				message={
					"stage": "Creating Purchase Order Schedule",
					"progress": round(float(idx) * 100 / total_records, 2),
					"description": po_number
				},
				user=frappe.session.user
			)

		except Exception as e:
			error_msg = f"Row {idx} failed: {e}"
			frappe.log_error(error_msg, "Purchase Order Upload Error")
			error_log.append(error_msg)

	# Second pass: submit docs
	total_to_submit = len(docs_to_submit)
	for idx, name in enumerate(docs_to_submit, start=1):
		try:
			doc = frappe.get_doc("Purchase Order Schedule", name)
			if doc.docstatus == 0:
				frappe.flags.ignore_permissions = True
				doc.submit()
			frappe.publish_realtime(
				event='purchase_order_upload_progress',
				message={
					"stage": "Submitting Schedule",
					"progress": round(float(idx) * 100 / total_to_submit, 2),
					"description": doc.purchase_order_number
				},
				user=frappe.session.user
			)
		except frappe.exceptions.PermissionError:
			error_msg = f"Permission error on submit: {name}"
			frappe.log_error(error_msg, "Purchase Order Upload Error")
			error_log.append(error_msg)
		except Exception as e:
			error_msg = f"Submit failed for {name}: {e}"
			frappe.log_error(error_msg, "Purchase Order Upload Error")
			error_log.append(error_msg)
		finally:
			frappe.flags.ignore_permissions = False

	# Third pass: update related Purchase Orders
	
	distinct_po_numbers = list(set(distinct_po_numbers))
	total_po = len(distinct_po_numbers)
	for idx, purchase_order_number in enumerate(distinct_po_numbers, start=1):
		try:
			order_open = frappe.get_doc("Purchase Open Order", {"purchase_order": purchase_order_number})
			order_open.disable_update_items = 0
			order_open.save(ignore_permissions=True)

			frappe.publish_realtime(
				event='purchase_order_upload_progress',
				message={
					"stage": "Updating Purchase Order",
					"progress": round(float(idx) * 100 / total_po, 2),
					"description": purchase_order_number
				},
				user=frappe.session.user
			)

		except Exception as e:
			error_msg = f"Update failed for PO {purchase_order_number}: {e}"
			frappe.log_error(error_msg, "Purchase Order Upload Error")
			error_log.append(error_msg)

	# Final status update
	if error_log:
		frappe.log_error("Error", "\n".join(error_log))
		frappe.publish_realtime(
			event='purchase_order_upload_progress',
			message={
				"stage": "Completed with Errors",
				"progress": 100,
				"errors": error_log
			},
			user=frappe.session.user
		)
	else:
		frappe.publish_realtime(
			event='purchase_order_upload_progress',
			message={
				"stage": "Completed Successfully",
				"progress": 100,
				"description": "All records processed."
			},
			user=frappe.session.user
		)
	now = frappe.utils.now_datetime().strftime("%d-%m-%Y %H:%M:%S")
	response_data = f"""
	<div style="
		max-width: 420px;
		margin: 20px auto;
		background: #ffffff;
		border-radius: 12px;
		padding: 20px 24px;
		box-shadow: 0 4px 14px rgba(0,0,0,0.08);
		font-family: Inter, sans-serif;
		border: 1px solid #e7e7e7;
	">
		<div style="font-size: 18px; font-weight: 600; color: #333;">
			Upload Summary
		</div>

		<div style="margin-top: 6px; font-size: 13px; color: #666;">
			{now}
		</div>

		<hr style="margin: 12px 0; border: 0; border-top: 1px solid #eee;">

		<ul style="margin: 0; padding-left: 20px; line-height: 1.8;">
			<li style="color:#4caf50">New documents created: {new_count}</li>
			<li style="color:#2196f3">Updated in existing document: {update_count}</li>
			<li style="color:#999">Skipped documents: {skipped_count}</li>
		</ul>
	</div>
	"""
	frappe.db.set_single_value("Purchase Order Schedule Settings", "response_data", response_data)
	frappe.clear_cache()
	frappe.publish_realtime("po_upload_summary", {
		"new": new_count,
		"update": update_count,
		"skip": skipped_count
	})
	return True



@frappe.whitelist()
def template_sheet():
	args = frappe.local.form_dict
	filename = args.name
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["Supplier Code","Purchase Order Number","Item Code","Schedule Period (month)","Schedule Qty"])
	ws.append(["","","","Example: Apr, May, Jun, Jul, Aug, Sep, ...",""])
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def get_data(file):
	data = """<table class=table table-bordered >
	<tr><td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>S.NO</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Supplier Code</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Supplier Name</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Purchase Order Number</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Code</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Name</b></center></td>

	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Schedule Period (in month)</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Schedule Qty</b></center></td>
	</tr>"""
	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])  
	i = 1  
	for pp in pps:
		if pp[0] != 'Supplier Code':
			
			sup_code = pp[0]
			sup_name = frappe.db.get_value("Supplier",pp[0],'supplier_name')
			po_number = pp[1]
			item = pp[2]
			item_name =frappe.db.get_value("Item",pp[2],'item_name')
			sch_date = pp[3]
			s_qty = pp[4]
			
			data += """
			
			<tr>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px; text-align: center;">%s</td>
			
		
			</tr>"""%(i,sup_code or '',sup_name or '',po_number or '',item or '',item_name or '',sch_date or '',s_qty or '')
			i += 1
	return data




@frappe.whitelist()
def enqueue_upload_validate_po(file):
	from frappe.utils.file_manager import get_file
	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])
	pps = [pp for pp in pps if pp and pp[0] != "Supplier Code"]

	if len(pps) > 500:
		return "<b>Upload supports only up to 500 rows</b>"
	error_logs = validate_attached_file_return_errors(file[0], pps)

	if error_logs:
		html = '''
		<div style="max-height: 300px; overflow-y: auto; border:1px solid #ddd; padding:10px; border-radius:5px; background:#f9f9f9;">
			<table class="table table-striped table-hover table-bordered">
				<thead style="background:#fd7e14; color:white;">
					<tr>
						<th style="width:60px;text-align:center">Row</th>
						<th style="text-align:center">Error</th>
					</tr>
				</thead>
				<tbody>
		'''

		for idx, err in enumerate(error_logs, start=1):
			html += f'''
			<tr>
				<td style="font-weight:bold; color:#dc3545;">{idx}</td>
				<td><i class="fa fa-exclamation-triangle" style="color:#dc3545; margin-right:5px;"></i>{err}</td>
			</tr>
			'''

		html += '''
				</tbody>
			</table>
		</div>
		'''

		return html




def validate_attached_file_return_errors(file, records):
	import frappe
	from frappe.utils.file_manager import remove_file
	error_logs = []
	valid_records = []
	allowed_months = {"JAN", "FEB", "MAR", "APR", "MAY", "JUN",
					  "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"}

	for idx, pp in enumerate(records, start=1):
		row_errors = []  # collect all errors for this row

		try:
			supplier_code, po_number, item, schedule_month = pp[0], pp[1], str(pp[2]), (pp[3] or "").strip()
			supplier_no = frappe.db.get_value("Supplier", {"supplier_code": supplier_code}, "name")
			frappe.errprint([supplier_code, supplier_no])
			s_qty = pp[4]

			if not supplier_no or not frappe.db.exists("Supplier", supplier_no):
				row_errors.append(f"Supplier with Supplier Code' {supplier_code}' does not exist.")

			if not po_number or not frappe.db.exists("Purchase Order", {"name": po_number, "docstatus": 1}):
				row_errors.append(f"Purchase Order '{po_number}' does not exist.")
			else:
				po = frappe.get_doc("Purchase Order", {"name": po_number, "docstatus": 1})
				if po.custom_order_type != "Open":
					row_errors.append(f"Purchase Order '{po_number}' is not of type 'Open'.")

				po_item = next((i for i in po.items if i.item_code == item), None)
				if not po_item:
					row_errors.append(f"Item '{item}' is not part of Purchase Order '{po_number}'.")

			try:
				if isinstance(s_qty, float) or ('.' in str(s_qty)):
					raise ValueError
				s_qty = int(s_qty)
				if s_qty < 0:
					raise ValueError
			except:
				row_errors.append(f"Schedule Qty '{s_qty}' must be a positive integer (no decimals).")

			schedule_month = schedule_month.upper()
			if schedule_month not in allowed_months:
				row_errors.append(f"Invalid Schedule Month '{schedule_month}'. Use Jan/Feb/... format.")

			exists = frappe.db.exists("Purchase Order Schedule", {
				"purchase_order_number": po_number,
				"schedule_month": schedule_month,
				"item_code": item,
				"docstatus": 1
			})
			if exists:
				exists_val = frappe.db.get_value("Purchase Order Schedule", {"purchase_order_number": po_number,"schedule_month": schedule_month,"item_code": item,"docstatus": 1},"received_qty")
				if exists_val and exists_val > s_qty:
					row_errors.append(f"While Revising the Schedule for {po_number}, {item}, {schedule_month}, the schedule qty{s_qty} is less than Received qty{exists_val}.")

			# if no errors, add to valid records
			if not row_errors:
				valid_records.append([supplier_no, po_number, item, schedule_month, s_qty])
			else:
				# Combine multiple errors into a single string (separated by <br>)
				error_logs.append( f"<b>Row {idx}</b>: " + "<br>".join(row_errors))

		except Exception as e:
			error_logs.append(f"Unexpected error - {str(e)}")
			frappe.log_error(f"Row {idx}: {str(e)}", "Purchase Order Upload Error")

	return error_logs
